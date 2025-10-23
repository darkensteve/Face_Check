from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
import sqlite3
from datetime import datetime
import os
import time
import secrets
import warnings

# Suppress known deprecation warning from face_recognition_models
warnings.filterwarnings('ignore', category=UserWarning, module='face_recognition_models')

# Check for required packages
try:
    import bcrypt 
    BCRYPT_AVAILABLE = True
except ImportError:
    print("Warning: bcrypt not available. Passwords will not be secure!")
    BCRYPT_AVAILABLE = False

try:
    import cv2 
    import numpy as np 
    import face_recognition 
    FACE_RECOGNITION_AVAILABLE = True
    print("Face recognition libraries loaded successfully")
except ImportError as e:
    print(f"Face recognition not available: {e}")
    print("Install with: pip install face-recognition opencv-contrib-python")
    FACE_RECOGNITION_AVAILABLE = False

# Import anti-spoofing module
try:
    from anti_spoofing import AntiSpoofingDetector, anti_spoofing_detector
    ANTI_SPOOFING_AVAILABLE = True
    print("Anti-spoofing module loaded successfully")
except ImportError as e:
    print(f"Anti-spoofing not available: {e}")
    ANTI_SPOOFING_AVAILABLE = False

app = Flask(__name__)
# Generate secure secret key from environment or create new one
app.secret_key = os.environ.get('SECRET_KEY') or secrets.token_hex(32)

# Session configuration for security
app.config.update(
    SESSION_COOKIE_SECURE=False,  # Set to True in production with HTTPS
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=7200,  # 2 hours session timeout (increased from 1 hour)
    SESSION_REFRESH_EACH_REQUEST=True  # Refresh session on each request
)

# Rate limiting for login attempts
login_attempts = {}

def is_rate_limited(ip_address, max_attempts=5, window_minutes=15):
    """Simple rate limiting for login attempts"""
    current_time = time.time()
    window_seconds = window_minutes * 60
    
    if ip_address not in login_attempts:
        login_attempts[ip_address] = []
    
    # Clean old attempts outside the window
    login_attempts[ip_address] = [
        attempt_time for attempt_time in login_attempts[ip_address]
        if current_time - attempt_time < window_seconds
    ]
    
    # Check if over limit
    if len(login_attempts[ip_address]) >= max_attempts:
        return True
    
    return False

def record_login_attempt(ip_address):
    """Record a failed login attempt"""
    if ip_address not in login_attempts:
        login_attempts[ip_address] = []
    login_attempts[ip_address].append(time.time())

# Database connection with better error handling
def get_db_connection():
    """Get database connection with error handling and timeout"""
    try:
        conn = sqlite3.connect('facecheck.db', timeout=30.0)
        conn.row_factory = sqlite3.Row
        # Enable foreign key constraints
        conn.execute('PRAGMA foreign_keys = ON')
        # Ensure expected schema exists (idempotent)
        try:
            cur = conn.cursor()
            cur.execute("PRAGMA table_info(faculty)")
            columns = [row[1] for row in cur.fetchall()]
            if 'attendance_image' not in columns:
                cur.execute("ALTER TABLE faculty ADD COLUMN attendance_image VARCHAR(255)")
                conn.commit()
        except Exception:
            # Ignore if PRAGMA/ALTER not applicable; app may still function without this column
            pass
        return conn
    except sqlite3.Error as e:
        print(f"Database connection error: {e}")
        raise


# Helper to safely format database datetime values which may be stored/returned
# as strings (most common) or as datetime objects. Prevents AttributeError when
# code calls .strftime on a string.
def safe_strftime(value, fmt):
    from datetime import datetime as _dt
    if not value:
        return ''
    # If it's already a datetime object
    if isinstance(value, _dt):
        return value.strftime(fmt)
    # If it's a string, try common formats
    if isinstance(value, str):
        for f in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try:
                parsed = _dt.strptime(value, f)
                return parsed.strftime(fmt)
            except Exception:
                continue
        # Fallback: return the raw string when parsing fails
        return value
    # Unknown type: convert to str
    return str(value)

# Authentication functions
def hash_password(password):
    """Hash password using bcrypt for secure storage"""
    if BCRYPT_AVAILABLE:
        import bcrypt 
        return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    else:
        # Fallback: NOT SECURE - only for development
        print("WARNING: Using insecure password storage!")
        return f"INSECURE_{password}"

def verify_password(stored_hash, password):
    """Verify password against stored hash"""
    if BCRYPT_AVAILABLE and not stored_hash.startswith('INSECURE_'):
        import bcrypt 
        try:
            return bcrypt.checkpw(password.encode('utf-8'), stored_hash.encode('utf-8'))
        except ValueError:
            # Handle legacy passwords
            return stored_hash == password
    else:
        # Fallback for insecure storage or legacy passwords
        return stored_hash == password or stored_hash == f"INSECURE_{password}"

def validate_input(data, field_type='text', min_len=1, max_len=255):
    """Validate and sanitize input data"""
    if not data or not isinstance(data, str):
        return False, "Invalid input"
    
    data = data.strip()
    
    if len(data) < min_len or len(data) > max_len:
        return False, f"Length must be between {min_len} and {max_len} characters"
    
    if field_type == 'idno':
        # Only allow alphanumeric characters for ID numbers
        if not data.replace('-', '').replace('_', '').isalnum():
            return False, "ID number can only contain letters, numbers, hyphens, and underscores"
    elif field_type == 'name':
        # Only allow letters, spaces, and common name characters
        import re
        if not re.match(r'^[a-zA-Z\s\.\-\']+$', data):
            return False, "Names can only contain letters, spaces, periods, hyphens, and apostrophes"
    elif field_type == 'role':
        if data not in ['admin', 'faculty', 'student']:
            return False, "Invalid role"
    
    return True, data

def authenticate_user(idno, password):
    """Authenticate user with secure password checking"""
    # Validate inputs
    is_valid_id, idno = validate_input(idno, 'idno', 1, 20)
    if not is_valid_id:
        return None
    
    is_valid_pass, password = validate_input(password, 'text', 1, 255)
    if not is_valid_pass:
        return None
    
    conn = get_db_connection()
    try:
        user = conn.execute(
            'SELECT * FROM user WHERE idno = ? AND is_active = 1',
            (idno,)
        ).fetchone()
        
        if user and verify_password(user['password'], password):
            return user
    except Exception as e:
        print(f"Authentication error: {e}")
    finally:
        conn.close()
    
    return None

def get_user_info(user_id):
    conn = get_db_connection()
    user = conn.execute(
        'SELECT u.*, d.dept_name FROM user u LEFT JOIN department d ON u.dept_id = d.dept_id WHERE u.user_id = ?',
        (user_id,)
    ).fetchone()
    conn.close()
    return user

def get_dashboard_stats():
    conn = get_db_connection()
    
    # Get total students
    total_students = conn.execute('SELECT COUNT(*) FROM student').fetchone()[0]
    
    # Get today's attendance
    today = datetime.now().strftime('%Y-%m-%d')
    today_attendance = conn.execute(
        'SELECT COUNT(*) FROM attendance WHERE DATE(attendance_date) = ?',
        (today,)
    ).fetchone()[0]
    
    # Get attendance rate
    attendance_rate = (today_attendance / total_students * 100) if total_students > 0 else 0
    
    # Get recent attendance
    recent_attendance = conn.execute('''
        SELECT a.attendance_date, u.firstname, u.lastname, a.attendance_status
        FROM attendance a
        JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
        JOIN student s ON sc.student_id = s.student_id
        JOIN user u ON s.user_id = u.user_id
        ORDER BY a.attendance_date DESC
        LIMIT 5
    ''').fetchall()
    
    conn.close()
    return {
        'total_students': total_students,
        'today_attendance': today_attendance,
        'attendance_rate': round(attendance_rate, 1),
        'recent_attendance': recent_attendance
    }

# Routes
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'unknown'))
        
        # Check rate limiting
        if is_rate_limited(client_ip):
            flash('Too many login attempts. Please try again in 15 minutes.', 'error')
            return render_template('login.html')
        
        idno = request.form.get('idno')
        password = request.form.get('password')
        
        if not idno or not password:
            flash('Please fill in all fields', 'error')
            record_login_attempt(client_ip)
            return render_template('login.html')
        
        user = authenticate_user(idno, password)
        if user:
            # Successful login - clear any rate limiting for this IP
            if client_ip in login_attempts:
                del login_attempts[client_ip]
            
            session.permanent = True
            session['user_id'] = user['user_id']
            session['idno'] = user['idno']
            session['role'] = user['role']
            session['firstname'] = user['firstname']
            session['lastname'] = user['lastname']
            
            if user['role'] == 'admin':
                return redirect(url_for('dashboard'))
            elif user['role'] == 'student':
                return redirect(url_for('student_dashboard'))
            elif user['role'] == 'faculty':
                return redirect(url_for('faculty_dashboard'))
        else:
            flash('Invalid credentials', 'error')
            record_login_attempt(client_ip)
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    stats = get_dashboard_stats()
    return render_template('dashboard.html', 
                         total_students=stats['total_students'],
                         today_attendance=stats['today_attendance'],
                         attendance_rate=stats['attendance_rate'],
                         recent_attendance=stats['recent_attendance'])

# User Management Routes
@app.route('/admin/users')
def admin_users():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    users = conn.execute('''
        SELECT u.*, d.dept_name, 
               CASE WHEN s.student_id IS NOT NULL THEN 'Student' 
                    WHEN f.faculty_id IS NOT NULL THEN 'Faculty'
                    ELSE 'Admin' END as user_type
        FROM user u
        LEFT JOIN department d ON u.dept_id = d.dept_id
        LEFT JOIN student s ON u.user_id = s.user_id
        LEFT JOIN faculty f ON u.user_id = f.user_id
        ORDER BY u.created_at DESC
    ''').fetchall()
    
    departments = conn.execute('SELECT * FROM department ORDER BY dept_name').fetchall()
    courses = conn.execute('SELECT * FROM course ORDER BY course_name').fetchall()
    
    conn.close()
    return render_template('admin_users.html', users=users, departments=departments, courses=courses)

@app.route('/admin/users/create', methods=['POST'])
def create_user():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        idno = request.form.get('idno')
        firstname = request.form.get('firstname')
        lastname = request.form.get('lastname')
        role = request.form.get('role')
        password = request.form.get('password')
        dept_id = request.form.get('dept_id')
        year_level = request.form.get('year_level')
        course_id = request.form.get('course_id')
        position = request.form.get('position')
        
        
        # Set default password if none provided (use ID number)
        if not password:
            password = idno  # Use ID number as default password
        
        # Validate required fields
        if not all([idno, firstname, lastname, role]):
            flash('Please fill in all required fields', 'error')
            return redirect(url_for('admin_users'))
        
        # Validate each field
        is_valid_id, validated_idno = validate_input(idno, 'idno', 1, 20)
        if not is_valid_id:
            flash(f'Invalid ID number: {validated_idno}', 'error')
            return redirect(url_for('admin_users'))
        
        is_valid_fname, validated_fname = validate_input(firstname, 'name', 1, 50)
        if not is_valid_fname:
            flash(f'Invalid first name: {validated_fname}', 'error')
            return redirect(url_for('admin_users'))
        
        is_valid_lname, validated_lname = validate_input(lastname, 'name', 1, 50)
        if not is_valid_lname:
            flash(f'Invalid last name: {validated_lname}', 'error')
            return redirect(url_for('admin_users'))
        
        is_valid_role, validated_role = validate_input(role, 'role')
        if not is_valid_role:
            flash(f'Invalid role: {validated_role}', 'error')
            return redirect(url_for('admin_users'))
        
        # Validate password strength
        if len(password) < 8:
            flash('Password must be at least 8 characters long', 'error')
            return redirect(url_for('admin_users'))
        
        # Validate department ID if provided
        if dept_id:
            try:
                dept_id = int(dept_id)
            except ValueError:
                flash('Invalid department ID', 'error')
                return redirect(url_for('admin_users'))
        
        # Validate course ID for students
        if validated_role == 'student' and course_id:
            try:
                course_id = int(course_id)
            except ValueError:
                flash('Invalid course ID', 'error')
                return redirect(url_for('admin_users'))
        
        # Validate password strength (minimum 3 characters for ID numbers)
        if len(password) < 3:
            flash('Password must be at least 3 characters long', 'error')
            return redirect(url_for('admin_users'))
        
        # Validate department ID if provided
        if dept_id:
            try:
                dept_id = int(dept_id)
            except ValueError:
                flash('Invalid department ID', 'error')
                return redirect(url_for('admin_users'))
        
        # Validate course ID for students
        if validated_role == 'student' and course_id:
            try:
                course_id = int(course_id)
            except ValueError:
                flash('Invalid course ID', 'error')
                return redirect(url_for('admin_users'))
        
        # Password is already set from form data or default
        
        conn = get_db_connection()
        
        # Check if user already exists
        existing_user = conn.execute('SELECT idno FROM user WHERE idno = ?', (validated_idno,)).fetchone()
        if existing_user:
            flash('User ID already exists', 'error')
            conn.close()
            return redirect(url_for('admin_users'))
        
        # Hash password before storing
        hashed_password = hash_password(password)
        
        # Insert user
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO user (idno, firstname, lastname, role, password, dept_id)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (validated_idno, validated_fname, validated_lname, validated_role, hashed_password, dept_id if dept_id else None))
        
        user_id = cursor.lastrowid
        
        # Insert role-specific data
        if validated_role == 'student' and course_id:
            if not year_level:
                flash('Year level is required for students', 'error')
                conn.rollback()
                conn.close()
                return redirect(url_for('admin_users'))
            
            cursor.execute('''
                INSERT INTO student (year_level, course_id, user_id)
                VALUES (?, ?, ?)
            ''', (year_level, course_id, user_id))
        elif validated_role == 'faculty' and position:
            cursor.execute('''
                INSERT INTO faculty (position, user_id)
                VALUES (?, ?)
            ''', (position, user_id))
        
        conn.commit()
        conn.close()
        
        flash('User created successfully', 'success')
        
    except Exception as e:
        flash(f'Error creating user: {str(e)}', 'error')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
def edit_user(user_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        try:
            firstname = request.form.get('firstname')
            lastname = request.form.get('lastname')
            role = request.form.get('role')
            dept_id = request.form.get('dept_id')
            year_level = request.form.get('year_level')
            course_id = request.form.get('course_id')
            position = request.form.get('position')
            is_active = request.form.get('is_active', '0')
            
            if not all([firstname, lastname, role]):
                flash('Please fill in all required fields', 'error')
                conn.close()
                return redirect(url_for('edit_user', user_id=user_id))
            
            # Update user
            conn.execute('''
                UPDATE user SET firstname = ?, lastname = ?, role = ?, dept_id = ?, is_active = ?
                WHERE user_id = ?
            ''', (firstname, lastname, role, dept_id, is_active, user_id))
            
            # Update role-specific data
            if role == 'student' and course_id:
                # Check if student record exists
                student = conn.execute('SELECT student_id FROM student WHERE user_id = ?', (user_id,)).fetchone()
                if student:
                    conn.execute('''
                        UPDATE student SET year_level = ?, course_id = ?
                        WHERE user_id = ?
                    ''', (year_level, course_id, user_id))
                else:
                    conn.execute('''
                        INSERT INTO student (year_level, course_id, user_id)
                        VALUES (?, ?, ?)
                    ''', (year_level, course_id, user_id))
            elif role == 'faculty' and position:
                # Check if faculty record exists
                faculty = conn.execute('SELECT faculty_id FROM faculty WHERE user_id = ?', (user_id,)).fetchone()
                if faculty:
                    conn.execute('''
                        UPDATE faculty SET position = ?
                        WHERE user_id = ?
                    ''', (position, user_id))
                else:
                    conn.execute('''
                        INSERT INTO faculty (position, user_id)
                        VALUES (?, ?)
                    ''', (position, user_id))
            
            conn.commit()
            flash('User updated successfully', 'success')
            
        except Exception as e:
            flash(f'Error updating user: {str(e)}', 'error')
    
    # Get user data
    user = conn.execute('''
        SELECT u.*, d.dept_name, s.year_level, s.course_id, f.position
        FROM user u
        LEFT JOIN department d ON u.dept_id = d.dept_id
        LEFT JOIN student s ON u.user_id = s.user_id
        LEFT JOIN faculty f ON u.user_id = f.user_id
        WHERE u.user_id = ?
    ''', (user_id,)).fetchone()
    
    departments = conn.execute('SELECT * FROM department ORDER BY dept_name').fetchall()
    courses = conn.execute('SELECT * FROM course ORDER BY course_name').fetchall()
    
    conn.close()
    return render_template('edit_user.html', user=user, departments=departments, courses=courses)

@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
def reset_password(user_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        new_password = request.form.get('new_password')
        
        if not new_password:
            flash('Password cannot be empty', 'error')
            return redirect(url_for('admin_users'))
        
        conn = get_db_connection()
        conn.execute('UPDATE user SET password = ? WHERE user_id = ?', (new_password, user_id))
        conn.commit()
        conn.close()
        
        flash('Password reset successfully', 'success')
        
    except Exception as e:
        flash(f'Error resetting password: {str(e)}', 'error')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/toggle-status', methods=['POST'])
def toggle_user_status(user_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        conn = get_db_connection()
        
        # Get current status
        user = conn.execute('SELECT is_active FROM user WHERE user_id = ?', (user_id,)).fetchone()
        new_status = 0 if user['is_active'] else 1
        
        # Update status
        conn.execute('UPDATE user SET is_active = ? WHERE user_id = ?', (new_status, user_id))
        conn.commit()
        conn.close()
        
        status_text = 'activated' if new_status else 'deactivated'
        flash(f'User {status_text} successfully', 'success')
        
    except Exception as e:
        flash(f'Error updating user status: {str(e)}', 'error')
    
    return redirect(url_for('admin_users'))

# Class & Event Management Routes
@app.route('/admin/classes')
def admin_classes():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # Get all classes with faculty info
    classes = conn.execute('''
        SELECT c.*, u.firstname, u.lastname, d.dept_name
        FROM class c
        JOIN faculty f ON c.faculty_id = f.faculty_id
        JOIN user u ON f.user_id = u.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        ORDER BY c.class_name
    ''').fetchall()
    
    # Get all faculty for assignment
    faculty = conn.execute('''
        SELECT f.faculty_id, u.firstname, u.lastname, d.dept_name
        FROM faculty f
        JOIN user u ON f.user_id = u.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE u.is_active = 1
        ORDER BY u.firstname, u.lastname
    ''').fetchall()
    
    conn.close()
    return render_template('admin_classes.html', classes=classes, faculty=faculty)

@app.route('/admin/classes/create', methods=['GET', 'POST'])
def create_class():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            class_name = request.form.get('class_name')
            edpcode = request.form.get('edpcode')
            start_time = request.form.get('start_time')
            end_time = request.form.get('end_time')
            room = request.form.get('room')
            faculty_id = request.form.get('faculty_id')
            days = request.form.getlist('days')  # Multiple days can be selected
            
            if not all([class_name, edpcode, start_time, end_time, room, faculty_id]):
                flash('Please fill in all required fields', 'error')
                return redirect(url_for('admin_classes'))
            
            conn = get_db_connection()
            
            # Check if EDP code already exists
            existing_class = conn.execute('SELECT class_id FROM class WHERE edpcode = ?', (edpcode,)).fetchone()
            if existing_class:
                flash('EDP Code already exists', 'error')
                conn.close()
                return redirect(url_for('admin_classes'))
            
            # Insert class
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO class (class_name, edpcode, start_time, end_time, room, faculty_id)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (class_name, edpcode, start_time, end_time, room, faculty_id))
            
            class_id = cursor.lastrowid
            
            # Insert class days
            for day_id in days:
                cursor.execute('''
                    INSERT INTO class_days (class_id, day_id)
                    VALUES (?, ?)
                ''', (class_id, day_id))
            
            conn.commit()
            conn.close()
            
            flash('Class created successfully', 'success')
            return redirect(url_for('admin_classes'))
            
        except Exception as e:
            flash(f'Error creating class: {str(e)}', 'error')
    
    # GET request - show form
    conn = get_db_connection()
    faculty = conn.execute('''
        SELECT f.faculty_id, u.firstname, u.lastname, d.dept_name
        FROM faculty f
        JOIN user u ON f.user_id = u.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE u.is_active = 1
        ORDER BY u.firstname, u.lastname
    ''').fetchall()
    
    days = conn.execute('SELECT * FROM days ORDER BY day_id').fetchall()
    conn.close()
    
    return render_template('create_class.html', faculty=faculty, days=days)

@app.route('/admin/classes/<int:class_id>/edit', methods=['GET', 'POST'])
def edit_class(class_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        try:
            class_name = request.form.get('class_name')
            edpcode = request.form.get('edpcode')
            start_time = request.form.get('start_time')
            end_time = request.form.get('end_time')
            room = request.form.get('room')
            faculty_id = request.form.get('faculty_id')
            days = request.form.getlist('days')
            
            if not all([class_name, edpcode, start_time, end_time, room, faculty_id]):
                flash('Please fill in all required fields', 'error')
                conn.close()
                return redirect(url_for('edit_class', class_id=class_id))
            
            # Check if EDP code already exists (excluding current class)
            existing_class = conn.execute('SELECT class_id FROM class WHERE edpcode = ? AND class_id != ?', (edpcode, class_id)).fetchone()
            if existing_class:
                flash('EDP Code already exists', 'error')
                conn.close()
                return redirect(url_for('edit_class', class_id=class_id))
            
            # Update class
            conn.execute('''
                UPDATE class SET class_name = ?, edpcode = ?, start_time = ?, end_time = ?, room = ?, faculty_id = ?
                WHERE class_id = ?
            ''', (class_name, edpcode, start_time, end_time, room, faculty_id, class_id))
            
            # Update class days
            conn.execute('DELETE FROM class_days WHERE class_id = ?', (class_id,))
            for day_id in days:
                conn.execute('''
                    INSERT INTO class_days (class_id, day_id)
                    VALUES (?, ?)
                ''', (class_id, day_id))
            
            conn.commit()
            flash('Class updated successfully', 'success')
            
        except Exception as e:
            flash(f'Error updating class: {str(e)}', 'error')
    
    # Get class data
    class_info = conn.execute('''
        SELECT c.*, u.firstname, u.lastname, d.dept_name
        FROM class c
        JOIN faculty f ON c.faculty_id = f.faculty_id
        JOIN user u ON f.user_id = u.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE c.class_id = ?
    ''', (class_id,)).fetchone()
    
    # Get class days
    class_days = conn.execute('SELECT day_id FROM class_days WHERE class_id = ?', (class_id,)).fetchall()
    class_day_ids = [day['day_id'] for day in class_days]
    
    # Get faculty and days for form
    faculty = conn.execute('''
        SELECT f.faculty_id, u.firstname, u.lastname, d.dept_name
        FROM faculty f
        JOIN user u ON f.user_id = u.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE u.is_active = 1
        ORDER BY u.firstname, u.lastname
    ''').fetchall()
    
    days = conn.execute('SELECT * FROM days ORDER BY day_id').fetchall()
    
    conn.close()
    return render_template('edit_class.html', class_info=class_info, faculty=faculty, days=days, class_day_ids=class_day_ids)

@app.route('/admin/classes/<int:class_id>/students')
def class_students(class_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # Get class info
    class_info = conn.execute('''
        SELECT c.*, u.firstname, u.lastname
        FROM class c
        JOIN faculty f ON c.faculty_id = f.faculty_id
        JOIN user u ON f.user_id = u.user_id
        WHERE c.class_id = ?
    ''', (class_id,)).fetchone()
    
    # Get enrolled students
    enrolled_students = conn.execute('''
        SELECT u.idno, u.firstname, u.lastname, s.student_id, s.year_level, c.course_name, d.dept_name
        FROM student_class sc
        JOIN student s ON sc.student_id = s.student_id
        JOIN user u ON s.user_id = u.user_id
        LEFT JOIN course c ON s.course_id = c.course_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE sc.class_id = ?
        ORDER BY u.firstname, u.lastname
    ''', (class_id,)).fetchall()
    
    # Get available students (not enrolled in this class)
    available_students = conn.execute('''
        SELECT u.idno, u.firstname, u.lastname, s.student_id, s.year_level, c.course_name, d.dept_name
        FROM user u
        JOIN student s ON u.user_id = s.user_id
        LEFT JOIN course c ON s.course_id = c.course_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE u.is_active = 1 AND u.role = 'student'
        AND s.student_id NOT IN (
            SELECT sc.student_id FROM student_class sc WHERE sc.class_id = ?
        )
        ORDER BY u.firstname, u.lastname
    ''', (class_id,)).fetchall()
    
    conn.close()
    return render_template('class_students.html', 
                         class_info=class_info, 
                         enrolled_students=enrolled_students, 
                         available_students=available_students)

@app.route('/admin/classes/<int:class_id>/enroll', methods=['POST'])
def enroll_student_to_class(class_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        student_id = request.form.get('student_id')
        
        if not student_id:
            flash('Please select a student', 'error')
            return redirect(url_for('class_students', class_id=class_id))
        
        conn = get_db_connection()
        
        # Check if student is already enrolled
        existing_enrollment = conn.execute('''
            SELECT studentclass_id FROM student_class 
            WHERE class_id = ? AND student_id = ?
        ''', (class_id, student_id)).fetchone()
        
        if existing_enrollment:
            flash('Student is already enrolled in this class', 'error')
            conn.close()
            return redirect(url_for('class_students', class_id=class_id))
        
        # Enroll student
        conn.execute('''
            INSERT INTO student_class (class_id, student_id)
            VALUES (?, ?)
        ''', (class_id, student_id))
        
        conn.commit()
        conn.close()
        
        flash('Student enrolled successfully', 'success')
        
    except Exception as e:
        flash(f'Error enrolling student: {str(e)}', 'error')
    
    return redirect(url_for('class_students', class_id=class_id))

@app.route('/admin/classes/<int:class_id>/bulk-enroll', methods=['POST'])
def bulk_enroll_students(class_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        student_ids = request.form.getlist('student_ids')
        
        if not student_ids:
            flash('Please select at least one student', 'error')
            return redirect(url_for('class_students', class_id=class_id))
        
        conn = get_db_connection()
        enrolled_count = 0
        already_enrolled = 0
        
        for student_id in student_ids:
            # Check if student is already enrolled
            existing_enrollment = conn.execute('''
                SELECT studentclass_id FROM student_class 
                WHERE class_id = ? AND student_id = ?
            ''', (class_id, student_id)).fetchone()
            
            if not existing_enrollment:
                # Enroll student
                conn.execute('''
                    INSERT INTO student_class (class_id, student_id)
                    VALUES (?, ?)
                ''', (class_id, student_id))
                enrolled_count += 1
            else:
                already_enrolled += 1
        
        conn.commit()
        conn.close()
        
        if enrolled_count > 0:
            flash(f'Successfully enrolled {enrolled_count} student(s)', 'success')
        if already_enrolled > 0:
            flash(f'{already_enrolled} student(s) were already enrolled', 'warning')
        
    except Exception as e:
        flash(f'Error enrolling students: {str(e)}', 'error')
    
    return redirect(url_for('class_students', class_id=class_id))

@app.route('/admin/classes/<int:class_id>/bulk-unenroll', methods=['POST'])
def bulk_unenroll_students(class_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        student_ids = request.form.getlist('student_ids')
        
        if not student_ids:
            flash('Please select at least one student', 'error')
            return redirect(url_for('class_students', class_id=class_id))
        
        conn = get_db_connection()
        unenrolled_count = 0
        
        for student_id in student_ids:
            # Remove student from class
            conn.execute('''
                DELETE FROM student_class 
                WHERE class_id = ? AND student_id = ?
            ''', (class_id, student_id))
            unenrolled_count += 1
        
        conn.commit()
        conn.close()
        
        flash(f'Successfully removed {unenrolled_count} student(s) from the class', 'success')
        
    except Exception as e:
        flash(f'Error removing students: {str(e)}', 'error')
    
    return redirect(url_for('class_students', class_id=class_id))

@app.route('/admin/classes/<int:class_id>/unenroll/<int:student_id>', methods=['POST'])
def unenroll_student_from_class(class_id, student_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        conn = get_db_connection()
        
        # Remove student from class
        conn.execute('''
            DELETE FROM student_class 
            WHERE class_id = ? AND student_id = ?
        ''', (class_id, student_id))
        
        conn.commit()
        conn.close()
        
        flash('Student unenrolled successfully', 'success')
        
    except Exception as e:
        flash(f'Error unenrolling student: {str(e)}', 'error')
    
    return redirect(url_for('class_students', class_id=class_id))

# Event Management Routes
@app.route('/admin/events')
def admin_events():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # Get all events with faculty info
    events = conn.execute('''
        SELECT e.*, u.firstname, u.lastname, d.dept_name
        FROM event e
        JOIN faculty f ON e.faculty_id = f.faculty_id
        JOIN user u ON f.user_id = u.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        ORDER BY e.event_date DESC
    ''').fetchall()
    
    # Get all faculty for assignment
    faculty = conn.execute('''
        SELECT f.faculty_id, u.firstname, u.lastname, d.dept_name
        FROM faculty f
        JOIN user u ON f.user_id = u.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE u.is_active = 1
        ORDER BY u.firstname, u.lastname
    ''').fetchall()
    
    conn.close()
    return render_template('admin_events.html', events=events, faculty=faculty)

@app.route('/admin/events/create', methods=['GET', 'POST'])
def create_event():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            event_name = request.form.get('event_name')
            description = request.form.get('description')
            event_date = request.form.get('event_date')
            start_time = request.form.get('start_time')
            end_time = request.form.get('end_time')
            room = request.form.get('room')
            faculty_id = request.form.get('faculty_id')
            
            if not all([event_name, event_date, start_time, end_time, faculty_id]):
                flash('Please fill in all required fields', 'error')
                return redirect(url_for('admin_events'))
            
            conn = get_db_connection()
            
            # Insert event
            conn.execute('''
                INSERT INTO event (event_name, description, event_date, start_time, end_time, room, faculty_id)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (event_name, description, event_date, start_time, end_time, room, faculty_id))
            
            conn.commit()
            conn.close()
            
            flash('Event created successfully', 'success')
            return redirect(url_for('admin_events'))
            
        except Exception as e:
            flash(f'Error creating event: {str(e)}', 'error')
    
    # GET request - show form
    conn = get_db_connection()
    faculty = conn.execute('''
        SELECT f.faculty_id, u.firstname, u.lastname, d.dept_name
        FROM faculty f
        JOIN user u ON f.user_id = u.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE u.is_active = 1
        ORDER BY u.firstname, u.lastname
    ''').fetchall()
    conn.close()
    
    return render_template('create_event.html', faculty=faculty)

@app.route('/admin/events/<int:event_id>/edit', methods=['GET', 'POST'])
def edit_event(event_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        try:
            event_name = request.form.get('event_name')
            description = request.form.get('description')
            event_date = request.form.get('event_date')
            start_time = request.form.get('start_time')
            end_time = request.form.get('end_time')
            room = request.form.get('room')
            faculty_id = request.form.get('faculty_id')
            
            if not all([event_name, event_date, start_time, end_time, faculty_id]):
                flash('Please fill in all required fields', 'error')
                conn.close()
                return redirect(url_for('edit_event', event_id=event_id))
            
            # Update event
            conn.execute('''
                UPDATE event SET event_name = ?, description = ?, event_date = ?, start_time = ?, end_time = ?, room = ?, faculty_id = ?
                WHERE event_id = ?
            ''', (event_name, description, event_date, start_time, end_time, room, faculty_id, event_id))
            
            conn.commit()
            flash('Event updated successfully', 'success')
            
        except Exception as e:
            flash(f'Error updating event: {str(e)}', 'error')
    
    # Get event data
    event_info = conn.execute('''
        SELECT e.*, u.firstname, u.lastname, d.dept_name
        FROM event e
        JOIN faculty f ON e.faculty_id = f.faculty_id
        JOIN user u ON f.user_id = u.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE e.event_id = ?
    ''', (event_id,)).fetchone()
    
    # Get faculty for form
    faculty = conn.execute('''
        SELECT f.faculty_id, u.firstname, u.lastname, d.dept_name
        FROM faculty f
        JOIN user u ON f.user_id = u.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE u.is_active = 1
        ORDER BY u.firstname, u.lastname
    ''').fetchall()
    
    conn.close()
    return render_template('edit_event.html', event_info=event_info, faculty=faculty)

# Face Registration Route
@app.route('/register_face')
def register_face():
    if 'user_id' not in session or session['role'] != 'student':
        return redirect(url_for('login'))
    
    # Get student info for the registration process
    conn = get_db_connection()
    student = conn.execute('''
        SELECT u.idno, u.firstname, u.lastname
        FROM user u
        JOIN student s ON u.user_id = s.user_id
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    if not student:
        conn.close()
        flash('Student not found', 'error')
        return redirect(url_for('student_dashboard'))
    
    conn.close()
    
    return render_template('register_face.html', 
                         student_name=student['firstname'] + ' ' + student['lastname'],
                         student_id=student['idno'])

# Faculty Face Registration Route
@app.route('/faculty/register_face')
def faculty_register_face():
    if 'user_id' not in session or session['role'] != 'faculty':
        return redirect(url_for('login'))
    
    # Get faculty info for the registration process
    conn = get_db_connection()
    faculty = conn.execute('''
        SELECT u.idno, u.firstname, u.lastname
        FROM user u
        JOIN faculty f ON u.user_id = f.user_id
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    if not faculty:
        conn.close()
        flash('Faculty not found', 'error')
        return redirect(url_for('faculty_dashboard'))
    
    conn.close()
    
    return render_template('faculty/faculty_register_face.html', 
                         faculty_name=faculty['firstname'] + ' ' + faculty['lastname'],
                         faculty_id=faculty['idno'])

# Student Dashboard
@app.route('/student/dashboard')
def student_dashboard():
    if 'user_id' not in session or session['role'] != 'student':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # Get student info
    student = conn.execute('''
        SELECT u.*, s.student_id, s.year_level, s.attendance_image, c.course_name, d.dept_name
        FROM user u
        JOIN student s ON u.user_id = s.user_id
        LEFT JOIN course c ON s.course_id = c.course_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    # Get today's attendance
    today = datetime.now().strftime('%Y-%m-%d')
    today_attendance = conn.execute('''
        SELECT a.attendance_status, a.attendance_date
        FROM attendance a
        JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
        WHERE sc.student_id = ? AND DATE(a.attendance_date) = ?
    ''', (student['student_id'], today)).fetchall()
    
    # Get attendance history
    attendance_history = conn.execute('''
        SELECT a.attendance_date, a.attendance_status, cl.class_name
        FROM attendance a
        JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
        JOIN class cl ON sc.class_id = cl.class_id
        WHERE sc.student_id = ?
        ORDER BY a.attendance_date DESC
        LIMIT 10
    ''', (student['student_id'],)).fetchall()
    
    # Format attendance history for template
    formatted_history = []
    for record in attendance_history:
        date_str = ''
        time_str = ''
        if record['attendance_date']:
            try:
                if isinstance(record['attendance_date'], str):
                    # Format: '2025-09-25 08:44:50' -> extract date and time
                    if ' ' in record['attendance_date']:
                        date_str = record['attendance_date'].split(' ')[0]
                        time_str = record['attendance_date'].split(' ')[1]
                    else:
                        date_str = record['attendance_date']
                        time_str = ''
                else:
                    # If it's a datetime object, use strftime
                    date_str = record['attendance_date'].strftime('%Y-%m-%d')
                    time_str = record['attendance_date'].strftime('%H:%M:%S')
            except:
                date_str = str(record['attendance_date'])
                time_str = ''
        
        formatted_history.append({
            'date': date_str,
            'time': time_str,
            'status': record['attendance_status']
        })
    
    # Get today's attendance status (single record or None)
    today_status = today_attendance[0] if today_attendance else None
    
    # Check if student has registered their face (check database first, then file system)
    has_face_registered = False
    
    # Check if attendance_image is stored in database
    if student['attendance_image']:
        # Verify the file actually exists
        if os.path.exists(student['attendance_image']):
            has_face_registered = True
        else:
            # File doesn't exist, clear the database record
            conn.execute('UPDATE student SET attendance_image = NULL WHERE user_id = ?', (session['user_id'],))
            conn.commit()
    
    # Fallback: check file system directly (for backward compatibility)
    if not has_face_registered:
        face_image_path = f"known_faces/{student['idno']}.jpg"
        if os.path.exists(face_image_path):
            # Update database with the file path
            conn.execute('UPDATE student SET attendance_image = ? WHERE user_id = ?', (face_image_path, session['user_id']))
            conn.commit()
            has_face_registered = True
    
    conn.close()
    return render_template('student_dashboard.html', 
                         student_info=student, 
                         today_attendance=today_status,
                         attendance_history=formatted_history,
                         has_face_registered=has_face_registered)

# Faculty Dashboard
@app.route('/faculty/dashboard')
def faculty_dashboard():
    if 'user_id' not in session or session['role'] != 'faculty':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # Get faculty info
    faculty = conn.execute('''
        SELECT u.*, f.faculty_id, f.position, f.attendance_image, d.dept_name
        FROM user u
        JOIN faculty f ON u.user_id = f.user_id
        LEFT JOIN department d ON u.dept_id = d.dept_id
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    # Get faculty stats
    my_students = conn.execute('''
        SELECT COUNT(DISTINCT sc.student_id) as student_count
        FROM class c
        JOIN student_class sc ON c.class_id = sc.class_id
        WHERE c.faculty_id = ?
    ''', (faculty['faculty_id'],)).fetchone()
    
    # Get today's attendance for faculty's classes
    today = datetime.now().strftime('%Y-%m-%d')
    today_attendance = conn.execute('''
        SELECT a.attendance_date, u.firstname, u.lastname, a.attendance_status, cl.class_name
        FROM attendance a
        JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
        JOIN student s ON sc.student_id = s.student_id
        JOIN user u ON s.user_id = u.user_id
        JOIN class cl ON sc.class_id = cl.class_id
        WHERE cl.faculty_id = ? AND DATE(a.attendance_date) = ?
        ORDER BY a.attendance_date DESC
    ''', (faculty['faculty_id'], today)).fetchall()
    
    # Calculate stats
    stats = {
        'total_students': my_students['student_count'] if my_students else 0,
        'today_present': len(today_attendance),
        'attendance_rate': 0  # Will be calculated based on total students
    }
    
    # Format attendance data for template
    formatted_attendance = []
    for record in today_attendance:
        # Extract time from datetime string
        time_str = ''
        if record['attendance_date']:
            try:
                # If it's already a string, extract the time part
                if isinstance(record['attendance_date'], str):
                    # Format: '2025-09-25 08:44:50' -> extract '08:44:50'
                    time_str = record['attendance_date'].split(' ')[1] if ' ' in record['attendance_date'] else record['attendance_date']
                else:
                    # If it's a datetime object, use strftime
                    time_str = record['attendance_date'].strftime('%H:%M:%S')
            except:
                time_str = str(record['attendance_date'])
        
        formatted_attendance.append({
            'name': f"{record['firstname']} {record['lastname']}",
            'time': time_str,
            'status': record['attendance_status']
        })
    
    # Check if faculty has registered their face
    has_face_registered = False
    if faculty['attendance_image']:
        import os
        if os.path.exists(faculty['attendance_image']):
            has_face_registered = True
        else:
            # File doesn't exist, clear the database record
            conn.execute('UPDATE faculty SET attendance_image = NULL WHERE user_id = ?', (session['user_id'],))
            conn.commit()
    
    conn.close()
    return render_template('faculty/faculty_dashboard.html', 
                         faculty_info=faculty, 
                         stats=stats, 
                         today_attendance=formatted_attendance,
                         has_face_registered=has_face_registered)

# API Routes
@app.route('/api/register_face', methods=['POST'])
def api_register_face():
    if 'user_id' not in session or session['role'] != 'student':
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get the uploaded face image
        if 'face_image' not in request.files:
            return jsonify({'error': 'No face image provided'}), 400
        
        face_file = request.files['face_image']
        if face_file.filename == '':
            return jsonify({'error': 'No face image selected'}), 400
        
        # Validate file type and size
        allowed_extensions = {'png', 'jpg', 'jpeg'}
        max_file_size = 5 * 1024 * 1024  # 5MB
        
        def allowed_file(filename):
            return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions
        
        if not allowed_file(face_file.filename):
            return jsonify({'error': 'Invalid file type. Only PNG, JPG, and JPEG are allowed'}), 400
        
        # Check file size
        face_file.seek(0, os.SEEK_END)
        file_size = face_file.tell()
        face_file.seek(0)
        
        if file_size > max_file_size:
            return jsonify({'error': 'File too large. Maximum size is 5MB'}), 400
        
        if file_size < 1024:  # Minimum 1KB
            return jsonify({'error': 'File too small. Minimum size is 1KB'}), 400
        
        # Get student info
        conn = get_db_connection()
        student = conn.execute('''
            SELECT u.idno FROM user u
            JOIN student s ON u.user_id = s.user_id
            WHERE u.user_id = ?
        ''', (session['user_id'],)).fetchone()
        
        if not student:
            conn.close()
            return jsonify({'error': 'Student not found'}), 404
        
        # Check if face recognition is available
        if not FACE_RECOGNITION_AVAILABLE:
            return jsonify({'error': 'Face recognition system is not configured. Please install required packages: pip install face-recognition opencv-contrib-python'}), 503
        
        # Create known_faces directory with proper permissions
        os.makedirs('known_faces', mode=0o755, exist_ok=True)
        
        # Read and validate the uploaded image
        try:
            face_data = face_file.read()
            if len(face_data) == 0:
                return jsonify({'error': 'Empty file received'}), 400
                
            nparr = np.frombuffer(face_data, np.uint8)
            image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        except Exception as e:
            return jsonify({'error': 'Failed to read image data'}), 400
        
        if image is None:
            return jsonify({'error': 'Invalid image format or corrupted file'}), 400
        
        # Validate image dimensions
        height, width = image.shape[:2]
        if height < 100 or width < 100:
            return jsonify({'error': 'Image too small. Minimum resolution is 100x100 pixels'}), 400
        
        if height > 4000 or width > 4000:
            return jsonify({'error': 'Image too large. Maximum resolution is 4000x4000 pixels'}), 400
        
        # Convert BGR to RGB for face_recognition
        rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        
        # Detect faces using face_recognition library or fallback
        try:
            if FACE_RECOGNITION_AVAILABLE:
                face_locations = face_recognition.face_locations(rgb_image, model="hog")
                face_encodings = face_recognition.face_encodings(rgb_image, face_locations)
            else:
                # Use OpenCV fallback
                from opencv_face_detector import simple_detector
                faces = simple_detector.detect_face(image)
                if len(faces) == 0:
                    face_encodings = []
                elif len(faces) > 1:
                    return jsonify({'error': 'Multiple faces detected. Please ensure only your face is visible in the camera.'}), 400
                else:
                    face_encodings = [1]  # Dummy encoding to indicate face found
        except Exception as e:
            return jsonify({'error': f'Face detection failed: {str(e)}'}), 400
        
        if not face_encodings:
            return jsonify({'error': 'No face detected in the image. Please ensure your face is clearly visible and try again.'}), 400
        
        if len(face_encodings) > 1:
            return jsonify({'error': 'Multiple faces detected. Please ensure only your face is visible in the camera.'}), 400
        
        # Generate secure filename
        import uuid
        filename = f"{student['idno']}_{uuid.uuid4().hex[:8]}.jpg"
        face_path = os.path.join('known_faces', filename)
        
        # Resize image to standard size to reduce storage
        max_size = 800
        if max(height, width) > max_size:
            scale = max_size / max(height, width)
            new_width = int(width * scale)
            new_height = int(height * scale)
            image = cv2.resize(image, (new_width, new_height), interpolation=cv2.INTER_AREA)
        
        # Save the face image with proper error handling
        try:
            if not cv2.imwrite(face_path, image):
                return jsonify({'error': 'Failed to save face image'}), 500
        except Exception as e:
            return jsonify({'error': f'Failed to save face image: {str(e)}'}), 500
        
        # Update the student record with the attendance_image path
        try:
            conn.execute('''
                UPDATE student 
                SET attendance_image = ? 
                WHERE user_id = ?
            ''', (face_path, session['user_id']))
            
            conn.commit()
        except Exception as e:
            # Clean up the saved file if database update fails
            if os.path.exists(face_path):
                os.remove(face_path)
            return jsonify({'error': f'Failed to update student record: {str(e)}'}), 500
        finally:
            conn.close()
        
        return jsonify({'success': True, 'message': 'Face registered successfully'})
        
    except Exception as e:
        return jsonify({'error': f'Registration failed: {str(e)}'}), 500

@app.route('/api/faculty/register_face', methods=['POST'])
def api_faculty_register_face():
    if 'user_id' not in session or session['role'] != 'faculty':
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get the uploaded face image
        if 'face_image' not in request.files:
            return jsonify({'error': 'No face image provided'}), 400
        
        face_file = request.files['face_image']
        if face_file.filename == '':
            return jsonify({'error': 'No face image selected'}), 400
        
        # Get faculty info
        conn = get_db_connection()
        faculty = conn.execute('''
            SELECT u.idno FROM user u
            JOIN faculty f ON u.user_id = f.user_id
            WHERE u.user_id = ?
        ''', (session['user_id'],)).fetchone()
        
        if not faculty:
            conn.close()
            return jsonify({'error': 'Faculty not found'}), 404
        
        # Import required libraries for face detection
        import cv2
        import numpy as np
        import face_recognition
        import os
        
        # Create known_faces directory
        os.makedirs('known_faces', exist_ok=True)
        
        # Read the uploaded image
        face_data = face_file.read()
        nparr = np.frombuffer(face_data, np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if image is None:
            conn.close()
            return jsonify({'error': 'Invalid image format'}), 400
        
        # Convert BGR to RGB for face_recognition
        rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        
        # Detect faces using face_recognition library
        face_locations = face_recognition.face_locations(rgb_image)
        face_encodings = face_recognition.face_encodings(rgb_image, face_locations)
        
        if not face_encodings:
            conn.close()
            return jsonify({'error': 'No face detected in the image. Please ensure your face is clearly visible and try again.'}), 400
        
        if len(face_encodings) > 1:
            conn.close()
            return jsonify({'error': 'Multiple faces detected. Please ensure only your face is visible in the camera.'}), 400
        
        # Save the face image with faculty prefix to distinguish from students
        face_path = f"known_faces/faculty_{faculty['idno']}.jpg"
        cv2.imwrite(face_path, image)
        
        # Add attendance_image column to faculty table if it doesn't exist
        try:
            conn.execute('ALTER TABLE faculty ADD COLUMN attendance_image VARCHAR(255)')
        except:
            pass  # Column already exists
        
        # Update the faculty record with the attendance_image path
        conn.execute('''
            UPDATE faculty 
            SET attendance_image = ? 
            WHERE user_id = ?
        ''', (face_path, session['user_id']))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Face registered successfully'})
        
    except ImportError as e:
        return jsonify({'error': 'Face recognition libraries not installed. Please contact administrator.'}), 500
    except Exception as e:
        return jsonify({'error': f'Registration failed: {str(e)}'}), 500

@app.route('/api/students')
def api_students():
    conn = get_db_connection()
    students = conn.execute('''
        SELECT u.idno, u.firstname, u.lastname, s.year_level, c.course_name
        FROM user u
        JOIN student s ON u.user_id = s.user_id
        LEFT JOIN course c ON s.course_id = c.course_id
        WHERE u.role = 'student' AND u.is_active = 1
    ''').fetchall()
    conn.close()
    
    return jsonify([dict(student) for student in students])

def calculate_ear(eye_landmarks):
    """Calculate Eye Aspect Ratio (EAR) for blink detection"""
    import numpy as np 
    
    # Convert to numpy array
    eye = np.array(eye_landmarks)
    
    # Calculate distances
    A = np.linalg.norm(eye[1] - eye[5])  # Vertical distance 1
    B = np.linalg.norm(eye[2] - eye[4])  # Vertical distance 2
    C = np.linalg.norm(eye[0] - eye[3])    # Horizontal distance
    
    # Calculate EAR
    ear = (A + B) / (2.0 * C)
    return ear

def process_face_recognition(image_path):
    """Process face recognition using the same logic as face_recog_test.py"""
    try:
        # Check if face recognition is available
        if not FACE_RECOGNITION_AVAILABLE:
            print("Using OpenCV fallback for face recognition")
            try:
                from opencv_face_detector import fallback_face_recognition
                return fallback_face_recognition(image_path)
            except ImportError as e:
                return {
                    'success': False,
                    'message': f'Face recognition system is not configured: {e}',
                    'student_id': 'Unknown',
                    'student_name': 'Unknown'
                }
        
        print(f"Processing image: {image_path}")
        
        # Load known faces from database
        conn = get_db_connection()
        students = conn.execute('''
            SELECT s.student_id, u.firstname, u.lastname, s.attendance_image
            FROM student s
            JOIN user u ON s.user_id = u.user_id
            WHERE s.attendance_image IS NOT NULL
        ''').fetchall()
        conn.close()
        
        print(f"Found {len(students)} registered students")
        
        if not students:
            return {
                'success': False,
                'message': 'No registered students found',
                'student_id': 'Unknown',
                'student_name': 'Unknown'
            }
        
        # Load the image
        image = cv2.imread(image_path)
        if image is None:
            print("Could not load image")
            return {
                'success': False,
                'message': 'Could not load image',
                'student_id': 'Unknown',
                'student_name': 'Unknown'
            }
        
        print(f"Image loaded: {image.shape}")
        
        # Convert BGR to RGB
        rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        
        # Find face locations
        print("Detecting faces...")
        face_locations = face_recognition.face_locations(rgb_image, number_of_times_to_upsample=1, model="hog")
        print(f"Found {len(face_locations)} face(s)")
        
        if not face_locations:
            return {
                'success': False,
                'message': 'No face detected',
                'student_id': 'Unknown',
                'student_name': 'Unknown'
            }
        
        # Get face encodings and landmarks
        print("Encoding faces...")
        face_encodings = face_recognition.face_encodings(rgb_image, face_locations)
        face_landmarks = face_recognition.face_landmarks(rgb_image, face_locations)
        print(f"Generated {len(face_encodings)} face encoding(s)")
        
        if not face_encodings:
            return {
                'success': False,
                'message': 'Could not encode face',
                'student_id': 'Unknown',
                'student_name': 'Unknown'
            }
        
        # Perform anti-spoofing check
        anti_spoofing_result = {'is_live': True, 'confidence': 1.0, 'details': 'Anti-spoofing disabled'}
        if ANTI_SPOOFING_AVAILABLE and face_landmarks:
            print("Performing anti-spoofing analysis...")
            anti_spoofing_result = anti_spoofing_detector.comprehensive_anti_spoofing_check(
                rgb_image, face_landmarks[0], face_locations[0]
            )
            print(f"Anti-spoofing result: {anti_spoofing_result['details']}")
            
            # Check if face passes anti-spoofing
            if not anti_spoofing_result['is_live']:
                return {
                    'success': False,
                    'message': f"Spoofing attempt detected! {anti_spoofing_result['details']}",
                    'student_id': 'SPOOFING_DETECTED',
                    'student_name': 'Spoofing Attempt',
                    'anti_spoofing': anti_spoofing_result
                }
        
        # Calculate real liveness detection metrics
        eye_ratio = 0.0
        nose_motion = 0.0
        
        if face_landmarks:
            landmarks = face_landmarks[0]  # Get landmarks for the first face
            
            # Calculate Eye Aspect Ratio (EAR) for blink detection
            if 'left_eye' in landmarks and 'right_eye' in landmarks:
                left_eye = landmarks['left_eye']
                right_eye = landmarks['right_eye']
                
                # Calculate EAR for both eyes
                left_ear = calculate_ear(left_eye)
                right_ear = calculate_ear(right_eye)
                eye_ratio = (left_ear + right_ear) / 2.0
                
                print(f"Eye Aspect Ratio: {eye_ratio}")
            
            # Calculate nose motion (simplified - using nose tip position)
            if 'nose_tip' in landmarks:
                nose_tip = landmarks['nose_tip']
                if len(nose_tip) > 0:
                    # Use nose tip position as motion indicator
                    nose_motion = len(nose_tip) * 2.0  # Simplified motion calculation
                    print(f"Nose motion: {nose_motion}")
        
        print(f"Liveness metrics - Eye ratio: {eye_ratio}, Nose motion: {nose_motion}")
        
        # Compare with known faces
        print("Comparing with known faces...")
        best_match = None
        best_distance = float('inf')
        
        for student in students:
            print(f"Checking student: {student['firstname']} {student['lastname']}")
            if not student['attendance_image'] or not os.path.exists(student['attendance_image']):
                print(f"  No valid image: {student['attendance_image']}")
                continue
                
            # Load known face
            known_image = cv2.imread(student['attendance_image'])
            if known_image is None:
                print(f"  Could not load image: {student['attendance_image']}")
                continue
                
            known_rgb = cv2.cvtColor(known_image, cv2.COLOR_BGR2RGB)
            known_encodings = face_recognition.face_encodings(known_rgb)
            
            if not known_encodings:
                print(f"  Could not encode known face")
                continue
            
            # Calculate distance
            distances = face_recognition.face_distance(known_encodings, face_encodings[0])
            min_distance = min(distances)
            print(f"  Distance: {min_distance}")
            
            if min_distance < best_distance:
                best_distance = min_distance
                best_match = student
        
        # Check if match is good enough (tolerance from face_recog_test.py)
        MATCH_TOLERANCE = 0.62
        print(f"Best match: {best_match['firstname'] if best_match else 'None'}")
        print(f"Best distance: {best_distance}")
        print(f"Tolerance: {MATCH_TOLERANCE}")
        
        # Get face location coordinates for drawing box
        # Use landmarks for more accurate face bounding box (like real-world systems)
        face_box = None
        if face_landmarks and len(face_landmarks) > 0:
            # Calculate precise bounding box from facial landmarks
            landmarks = face_landmarks[0]
            all_points = []
            
            # Collect all landmark points
            for feature_name in ['chin', 'left_eyebrow', 'right_eyebrow', 'nose_bridge', 
                                 'nose_tip', 'left_eye', 'right_eye', 'top_lip', 'bottom_lip']:
                if feature_name in landmarks:
                    all_points.extend(landmarks[feature_name])
            
            if all_points:
                # Find min and max coordinates
                all_x = [p[0] for p in all_points]
                all_y = [p[1] for p in all_points]
                
                min_x = min(all_x)
                max_x = max(all_x)
                min_y = min(all_y)
                max_y = max(all_y)
                
                # Add small margin for better visualization (10% padding)
                width = max_x - min_x
                height = max_y - min_y
                margin_x = int(width * 0.15)  # 15% horizontal margin
                margin_y = int(height * 0.20)  # 20% vertical margin (more for forehead/hair)
                
                # Apply margins with boundary checking
                min_x = max(0, min_x - margin_x)
                min_y = max(0, min_y - margin_y)
                max_x = min(rgb_image.shape[1], max_x + margin_x)
                max_y = min(rgb_image.shape[0], max_y + margin_y)
                
                face_box = {
                    'x': int(min_x),
                    'y': int(min_y),
                    'width': int(max_x - min_x),
                    'height': int(max_y - min_y)
                }
        elif face_locations:
            # Fallback to basic face_locations if landmarks not available
            top, right, bottom, left = face_locations[0]
            face_box = {
                'x': int(left),
                'y': int(top),
                'width': int(right - left),
                'height': int(bottom - top)
            }
        
        if best_match and best_distance <= MATCH_TOLERANCE:
            print("Match found!")
            confidence = int((1 - best_distance) * 100)  # Convert distance to confidence percentage
            return {
                'success': True,
                'student_id': int(best_match['student_id']),
                'student_name': f"{best_match['firstname']} {best_match['lastname']}",
                'distance': float(best_distance),
                'confidence': confidence,
                'face_box': face_box,  # Add face location for drawing box
                'eye_ratio': float(eye_ratio),  # Real eye aspect ratio
                'nose_motion': float(nose_motion),  # Real nose motion
                'anti_spoofing': {
                    'is_live': bool(anti_spoofing_result.get('is_live', True)),
                    'confidence': float(anti_spoofing_result.get('confidence', 1.0)),
                    'details': str(anti_spoofing_result.get('details', 'Anti-spoofing disabled'))
                }
            }
        else:
            print("No match found")
            return {
                'success': False,
                'message': 'No matching student found',
                'student_id': 'Unknown',
                'student_name': 'Unknown',
                'distance': float(best_distance if best_match else 999.0),  # Use 999.0 instead of float('inf')
                'confidence': 0,
                'face_box': face_box,  # Still provide face box even if no match
                'eye_ratio': float(eye_ratio),
                'nose_motion': float(nose_motion),
                'anti_spoofing': {
                    'is_live': bool(anti_spoofing_result.get('is_live', True)),
                    'confidence': float(anti_spoofing_result.get('confidence', 1.0)),
                    'details': str(anti_spoofing_result.get('details', 'Anti-spoofing disabled'))
                }
            }
            
    except Exception as e:
        print(f"Error in process_face_recognition: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'message': f'Recognition error: {str(e)}',
            'student_id': 'Unknown',
            'student_name': 'Unknown'
        }

@app.route('/api/attendance/detect', methods=['POST'])
def api_attendance_detect():
    """Face detection and recognition endpoint"""
    # Debug session info
    print(f"Session contents: {dict(session)}")
    print(f"Session has user_id: {'user_id' in session}")
    print(f"Session role: {session.get('role', 'NO ROLE')}")
    
    if 'user_id' not in session:
        print("ERROR: No user_id in session! User needs to login again.")
        return jsonify({
            'success': False, 
            'message': 'Your session has expired. Please logout and login again to continue.',
            'expired': True,
            'redirect': '/login'
        }), 401
    
    # Allow both faculty and admin to access this endpoint
    user_role = session.get('role')
    if user_role not in ['faculty', 'admin']:
        print(f"ERROR: Invalid role: {user_role}")
        return jsonify({
            'success': False, 
            'message': f'Only faculty and admin can access this feature. You are logged in as: {user_role}',
            'wrong_role': True,
            'current_role': user_role
        }), 401
    
    filepath = None
    try:
        # Get the uploaded image
        if 'image' not in request.files:
            return jsonify({'success': False, 'message': 'No image provided'}), 400
        
        file = request.files['image']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No image selected'}), 400
        
        # Save the image temporarily with secure filename
        import uuid
        safe_filename = f"temp_{session['user_id']}_{uuid.uuid4().hex[:8]}.jpg"
        filepath = os.path.join('temp', safe_filename)
        os.makedirs('temp', mode=0o755, exist_ok=True)
        
        # Validate file size before saving
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > 10 * 1024 * 1024:  # 10MB limit
            return jsonify({'success': False, 'message': 'File too large'}), 400
        
        file.save(filepath)
        
        # Check if face recognition is available
        if not FACE_RECOGNITION_AVAILABLE:
            return jsonify({
                'success': False, 
                'message': 'Face recognition system is not configured. Please install required packages.',
                'student_id': 'Unknown',
                'student_name': 'Unknown'
            }), 503
        
        # Process the image for face recognition
        print(f"Processing face recognition for image: {filepath}")
        result = process_face_recognition(filepath)
        print(f"Recognition result: {result}")
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error in api_attendance_detect: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500
    finally:
        # Always clean up temp file
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except Exception as e:
                print(f"Warning: Failed to clean up temp file {filepath}: {e}")

@app.route('/api/attendance/mark', methods=['POST'])
def api_attendance_mark():
    """Mark attendance for a recognized student"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized - Please login'}), 401
    
    # Allow both faculty and admin to access this endpoint
    if session.get('role') not in ['faculty', 'admin']:
        return jsonify({'success': False, 'message': 'Unauthorized - Faculty or Admin access required'}), 401
    
    try:
        data = request.get_json()
        print(f"Attendance mark request data: {data}")
        student_id = data.get('student_id')
        student_name = data.get('student_name')
        class_id = data.get('class_id')
        anti_spoofing_data = data.get('anti_spoofing', {})
        
        print(f"Student ID: {student_id}, Student Name: {student_name}, Class ID: {class_id}")
        
        if not student_id or not student_name or not class_id:
            return jsonify({'success': False, 'message': 'Missing student or class information'}), 400
        
        # Check for spoofing attempts
        if student_id == 'SPOOFING_DETECTED':
            print("⚠️ Spoofing attempt blocked!")
            return jsonify({
                'success': False, 
                'message': 'Spoofing attempt detected! Please try again with a live face.',
                'spoofing_detected': True
            }), 403
        
        # Validate anti-spoofing results if available
        if ANTI_SPOOFING_AVAILABLE and anti_spoofing_data:
            is_live = anti_spoofing_data.get('is_live', True)
            confidence = anti_spoofing_data.get('confidence', 1.0)
            
            if not is_live or confidence < 0.5:
                print(f"⚠️ Anti-spoofing failed: live={is_live}, confidence={confidence}")
                return jsonify({
                    'success': False,
                    'message': f'Anti-spoofing check failed. Confidence: {confidence:.1%}',
                    'spoofing_detected': True,
                    'anti_spoofing_details': anti_spoofing_data.get('details', 'Low confidence score')
                }), 403
        
        # Mark attendance in database
        conn = get_db_connection()
        print("Database connection established")
        
        # Get the correct studentclass_id for this student and class
        student_class = conn.execute('''
            SELECT sc.studentclass_id FROM student_class sc
            WHERE sc.student_id = ? AND sc.class_id = ?
        ''', (student_id, class_id)).fetchone()
        
        if not student_class:
            print(f"No student_class found for student_id: {student_id}")
            conn.close()
            return jsonify({'success': False, 'message': 'Student not enrolled in any class'})
        
        studentclass_id = student_class['studentclass_id']
        print(f"Found studentclass_id: {studentclass_id} for student_id: {student_id}")
        
        # Check if already marked today
        today = datetime.now().strftime('%Y-%m-%d')
        print(f"Checking for existing attendance on {today}")
        existing = conn.execute('''
            SELECT attendance_id FROM attendance 
            WHERE studentclass_id = ? AND DATE(attendance_date) = ?
        ''', (studentclass_id, today)).fetchone()
        
        if existing:
            print("Already marked today")
            conn.close()
            return jsonify({'success': False, 'message': 'Already marked today'})
        
        # Insert attendance record
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"Inserting attendance record: studentclass_id={studentclass_id}, time={current_time}")
        conn.execute('''
            INSERT INTO attendance (studentclass_id, attendance_date, attendance_status)
            VALUES (?, ?, 'present')
        ''', (studentclass_id, current_time))
        
        conn.commit()
        print("Attendance record inserted successfully")
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Attendance marked for {student_name}',
            'student_id': student_id,
            'student_name': student_name,
            'time': datetime.now().strftime('%H:%M:%S')
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/attendance', methods=['POST'])
def api_mark_attendance():
    data = request.get_json()
    student_id = data.get('student_id')
    status = data.get('status', 'present')
    
    if not student_id:
        return jsonify({'error': 'Student ID required'}), 400
    
    conn = get_db_connection()
    
    # Get student_class_id
    student_class = conn.execute('''
        SELECT sc.studentclass_id FROM student_class sc
        JOIN student s ON sc.student_id = s.student_id
        WHERE s.user_id = ?
    ''', (student_id,)).fetchone()
    
    if student_class:
        # Insert attendance record
        conn.execute('''
            INSERT INTO attendance (attendance_date, attendance_status, studentclass_id)
            VALUES (?, ?, ?)
        ''', (datetime.now(), status, student_class['studentclass_id']))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Attendance marked successfully'})
    else:
        conn.close()
        return jsonify({'error': 'Student not enrolled in any class'}), 400

@app.route('/api/attendance/today')
def api_today_attendance():
    today = datetime.now().strftime('%Y-%m-%d')
    class_id = request.args.get('class_id')
    
    if not class_id:
        return jsonify([])
    
    conn = get_db_connection()
    
    attendance = conn.execute('''
        SELECT s.student_id, u.firstname, u.lastname, a.attendance_status, a.attendance_date
        FROM attendance a
        JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
        JOIN student s ON sc.student_id = s.student_id
        JOIN user u ON s.user_id = u.user_id
        WHERE DATE(a.attendance_date) = ? AND sc.class_id = ?
        ORDER BY a.attendance_date DESC
    ''', (today, class_id)).fetchall()
    
    # Format the data for frontend
    formatted_attendance = []
    for record in attendance:
        # Extract time from datetime string
        time_str = ''
        if record['attendance_date']:
            try:
                if isinstance(record['attendance_date'], str):
                    time_str = record['attendance_date'].split(' ')[1] if ' ' in record['attendance_date'] else record['attendance_date']
                else:
                    time_str = record['attendance_date'].strftime('%H:%M:%S')
            except:
                time_str = str(record['attendance_date'])
        
        formatted_attendance.append({
            'student_id': record['student_id'],
            'student_name': f"{record['firstname']} {record['lastname']}",
            'time': time_str,
            'status': record['attendance_status']
        })
    
    conn.close()
    return jsonify(formatted_attendance)

@app.route('/api/faculty/classes')
def api_faculty_classes():
    """Get classes assigned to the current faculty member"""
    if 'user_id' not in session or session['role'] != 'faculty':
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = get_db_connection()
    
    # First get the faculty_id for the current user
    faculty = conn.execute('''
        SELECT f.faculty_id FROM faculty f
        WHERE f.user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    if not faculty:
        conn.close()
        return jsonify([])
    
    # Then get classes assigned to this faculty
    classes = conn.execute('''
        SELECT c.class_id, c.class_name, c.edpcode, c.start_time, c.end_time, c.room, c.faculty_id
        FROM class c
        WHERE c.faculty_id = ?
        ORDER BY c.class_name
    ''', (faculty['faculty_id'],)).fetchall()
    
    conn.close()
    return jsonify([dict(record) for record in classes])

@app.route('/admin/classes/<int:class_id>/delete', methods=['POST'])
def delete_class(class_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    try:
        # Check if class has enrolled students
        enrolled_students = conn.execute('''
            SELECT COUNT(*) as count FROM class_student WHERE class_id = ?
        ''', (class_id,)).fetchone()
        
        if enrolled_students['count'] > 0:
            flash('Cannot delete class with enrolled students. Please unenroll all students first.', 'error')
            conn.close()
            return redirect(url_for('admin_classes'))
        
        # Delete class days first
        conn.execute('DELETE FROM class_day WHERE class_id = ?', (class_id,))
        
        # Delete the class
        conn.execute('DELETE FROM class WHERE class_id = ?', (class_id,))
        
        conn.commit()
        flash('Class deleted successfully', 'success')
        
    except Exception as e:
        flash(f'Error deleting class: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_classes'))

@app.route('/admin/events/<int:event_id>/delete', methods=['POST'])
def delete_event(event_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    try:
        # Delete the event
        conn.execute('DELETE FROM event WHERE event_id = ?', (event_id,))
        
        conn.commit()
        flash('Event deleted successfully', 'success')
        
    except Exception as e:
        flash(f'Error deleting event: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_events'))

@app.route('/admin/attendance')
def admin_attendance():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # Get all attendance records with student and class information
    attendance_records = conn.execute('''
        SELECT 
            a.attendance_id,
            a.attendance_date,
            a.attendance_status,
            u.firstname,
            u.lastname,
            u.idno,
            c.class_name,
            c.edpcode,
            fu.firstname as faculty_firstname,
            fu.lastname as faculty_lastname
        FROM attendance a
        JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
        JOIN student s ON sc.student_id = s.student_id
        JOIN user u ON s.user_id = u.user_id
        JOIN class c ON sc.class_id = c.class_id
        JOIN faculty f ON c.faculty_id = f.faculty_id
        JOIN user fu ON f.user_id = fu.user_id
        ORDER BY a.attendance_date DESC, u.lastname, u.firstname
    ''').fetchall()
    
    # Get attendance statistics
    total_records = len(attendance_records)
    present_count = len([r for r in attendance_records if r['attendance_status'] == 'present'])
    absent_count = total_records - present_count
    attendance_rate = (present_count / total_records * 100) if total_records > 0 else 0
    
    conn.close()
    
    return render_template('admin_attendance.html', 
                         attendance_records=attendance_records,
                         total_records=total_records,
                         present_count=present_count,
                         absent_count=absent_count,
                         attendance_rate=round(attendance_rate, 1))

@app.route('/reports')
def admin_reports():
    """Render admin reports and analytics page"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    return render_template('admin_reports.html')

@app.route('/api/admin/reports/<report_type>')
def api_admin_reports(report_type):
    """Get report data for admin"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    try:
        conn = get_db_connection()
        
        if report_type == 'class':
            # Class attendance summary
            data = conn.execute('''
                SELECT 
                    DATE(a.attendance_date) as date,
                    c.class_name as name,
                    COUNT(CASE WHEN a.attendance_status = 'present' THEN 1 END) as present,
                    COUNT(CASE WHEN a.attendance_status = 'absent' THEN 1 END) as absent,
                    COUNT(CASE WHEN a.attendance_status = 'late' THEN 1 END) as late,
                    ROUND(COUNT(CASE WHEN a.attendance_status = 'present' THEN 1 END) * 100.0 / COUNT(*), 1) as rate
                FROM attendance a
                JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
                JOIN class c ON sc.class_id = c.class_id
                WHERE DATE(a.attendance_date) BETWEEN ? AND ?
                GROUP BY DATE(a.attendance_date), c.class_id
                ORDER BY date DESC
            ''', (date_from, date_to)).fetchall()
            
        elif report_type == 'event':
            # Event attendance summary
            data = conn.execute('''
                SELECT 
                    DATE(ea.attendance_time) as date,
                    e.event_name as name,
                    COUNT(CASE WHEN ea.status = 'present' THEN 1 END) as present,
                    COUNT(CASE WHEN ea.status = 'absent' THEN 1 END) as absent,
                    COUNT(CASE WHEN ea.status = 'late' THEN 1 END) as late,
                    ROUND(COUNT(CASE WHEN ea.status = 'present' THEN 1 END) * 100.0 / COUNT(*), 1) as rate
                FROM event_attendance ea
                JOIN event e ON ea.event_id = e.event_id
                WHERE DATE(ea.attendance_time) BETWEEN ? AND ?
                GROUP BY DATE(ea.attendance_time), e.event_id
                ORDER BY date DESC
            ''', (date_from, date_to)).fetchall()
            
        elif report_type == 'absence':
            # Absence patterns
            data = conn.execute('''
                SELECT 
                    u.lastname || ', ' || u.firstname as name,
                    DATE(a.attendance_date) as date,
                    0 as present,
                    COUNT(CASE WHEN a.attendance_status = 'absent' THEN 1 END) as absent,
                    COUNT(CASE WHEN a.attendance_status = 'late' THEN 1 END) as late,
                    0 as rate
                FROM attendance a
                JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
                JOIN student s ON sc.student_id = s.student_id
                JOIN user u ON s.user_id = u.user_id
                WHERE DATE(a.attendance_date) BETWEEN ? AND ?
                    AND a.attendance_status IN ('absent', 'late')
                GROUP BY u.user_id, DATE(a.attendance_date)
                ORDER BY date DESC, name
            ''', (date_from, date_to)).fetchall()
            
        elif report_type == 'monthly':
            # Monthly summary
            data = conn.execute('''
                SELECT 
                    strftime('%Y-%m', a.attendance_date) as date,
                    'Monthly Total' as name,
                    COUNT(CASE WHEN a.attendance_status = 'present' THEN 1 END) as present,
                    COUNT(CASE WHEN a.attendance_status = 'absent' THEN 1 END) as absent,
                    COUNT(CASE WHEN a.attendance_status = 'late' THEN 1 END) as late,
                    ROUND(COUNT(CASE WHEN a.attendance_status = 'present' THEN 1 END) * 100.0 / COUNT(*), 1) as rate
                FROM attendance a
                WHERE DATE(a.attendance_date) BETWEEN ? AND ?
                GROUP BY strftime('%Y-%m', a.attendance_date)
                ORDER BY date DESC
            ''', (date_from, date_to)).fetchall()
        else:
            return jsonify({'success': False, 'error': 'Invalid report type'}), 400
        
        # Convert to list of dicts
        details = []
        for row in data:
            details.append({
                'date': row['date'],
                'name': row['name'],
                'present': row['present'],
                'absent': row['absent'],
                'late': row['late'],
                'rate': row['rate']
            })
        
        # Calculate summary
        total_present = sum(d['present'] for d in details)
        total_absent = sum(d['absent'] for d in details)
        total_late = sum(d['late'] for d in details)
        total = total_present + total_absent + total_late
        attendance_rate = round(total_present * 100.0 / total, 1) if total > 0 else 0
        
        # Generate trend data
        trend_labels = []
        trend_present = []
        trend_absent = []
        trend_late = []
        
        for d in details[:10]:  # Last 10 entries for trend
            trend_labels.insert(0, d['date'])
            trend_present.insert(0, d['present'])
            trend_absent.insert(0, d['absent'])
            trend_late.insert(0, d['late'])
        
        conn.close()
        
        return jsonify({
            'success': True,
            'summary': {
                'total_present': total_present,
                'total_absent': total_absent,
                'total_late': total_late,
                'attendance_rate': attendance_rate
            },
            'trend': {
                'labels': trend_labels,
                'present': trend_present,
                'absent': trend_absent,
                'late': trend_late
            },
            'details': details
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/reports/export/<fmt>')
def export_reports(fmt):
    """Export reports in various formats"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    report_type = request.args.get('type', 'class')
    
    try:
        conn = get_db_connection()
        
        # Get data based on report type (reuse logic from api_admin_reports)
        if report_type == 'class':
            data = conn.execute('''
                SELECT 
                    DATE(a.attendance_date) as date,
                    c.class_name as name,
                    COUNT(CASE WHEN a.attendance_status = 'present' THEN 1 END) as present,
                    COUNT(CASE WHEN a.attendance_status = 'absent' THEN 1 END) as absent,
                    COUNT(CASE WHEN a.attendance_status = 'late' THEN 1 END) as late
                FROM attendance a
                JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
                JOIN class c ON sc.class_id = c.class_id
                WHERE DATE(a.attendance_date) BETWEEN ? AND ?
                GROUP BY DATE(a.attendance_date), c.class_id
                ORDER BY date DESC
            ''', (date_from, date_to)).fetchall()
        else:
            # Simple default for other types
            data = []
        
        conn.close()
        
        if fmt == 'csv':
            import io
            import csv
            
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['Date', 'Class/Event', 'Present', 'Absent', 'Late'])
            
            for row in data:
                writer.writerow([row['date'], row['name'], row['present'], row['absent'], row['late']])
            
            response = app.make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv'
            response.headers['Content-Disposition'] = f'attachment; filename=report_{report_type}_{date_from}_{date_to}.csv'
            return response
            
        else:
            flash('Export format not yet implemented', 'info')
            return redirect(url_for('admin_reports'))
            
    except Exception as e:
        flash(f'Export error: {str(e)}', 'error')
        return redirect(url_for('admin_reports'))


# Faculty Manage Students
@app.route('/manage_students')
def faculty_manage_students():
    if 'user_id' not in session or session['role'] != 'faculty':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # Get faculty info
    faculty = conn.execute('''
        SELECT f.faculty_id FROM faculty f 
        JOIN user u ON f.user_id = u.user_id 
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    if not faculty:
        conn.close()
        flash('Faculty record not found', 'error')
        return redirect(url_for('faculty_dashboard'))
    
    # Get classes assigned to this faculty
    classes = conn.execute('''
        SELECT c.class_id, c.class_name, c.edpcode
        FROM class c
        WHERE c.faculty_id = ?
        ORDER BY c.class_name
    ''', (faculty['faculty_id'],)).fetchall()
    
    # Get events assigned to this faculty
    events = conn.execute('''
        SELECT e.event_id, e.event_name, e.event_date
        FROM event e
        WHERE e.faculty_id = ?
        ORDER BY e.event_date DESC
    ''', (faculty['faculty_id'],)).fetchall()
    
    conn.close()
    return render_template('faculty/faculty_manage_students.html', classes=classes, events=events)

@app.route('/faculty/students/<path:selection>')
def faculty_get_students(selection):
    if 'user_id' not in session or session['role'] != 'faculty':
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = get_db_connection()
    
    # Get faculty info
    faculty = conn.execute('''
        SELECT f.faculty_id FROM faculty f 
        JOIN user u ON f.user_id = u.user_id 
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    if not faculty:
        conn.close()
        return jsonify({'error': 'Faculty record not found'}), 404
    
    students = []
    
    if selection.startswith('class_'):
        class_id = selection.replace('class_', '')
        # Get students enrolled in this class
        students = conn.execute('''
            SELECT u.user_id, u.idno, u.firstname, u.lastname, u.is_active,
                   s.year_level, c.course_name
            FROM student_class sc
            JOIN student s ON sc.student_id = s.student_id
            JOIN user u ON s.user_id = u.user_id
            LEFT JOIN course c ON s.course_id = c.course_id
            JOIN class cl ON sc.class_id = cl.class_id
            WHERE sc.class_id = ? AND cl.faculty_id = ?
            ORDER BY u.firstname, u.lastname
        ''', (class_id, faculty['faculty_id'])).fetchall()
        
    elif selection.startswith('event_'):
        event_id = selection.replace('event_', '')
        # Get students who have attended this event
        students = conn.execute('''
            SELECT DISTINCT u.user_id, u.idno, u.firstname, u.lastname, u.is_active,
                   s.year_level, c.course_name
            FROM event_attendance ea
            JOIN user u ON ea.user_id = u.user_id
            JOIN student s ON u.user_id = s.user_id
            LEFT JOIN course c ON s.course_id = c.course_id
            JOIN event e ON ea.event_id = e.event_id
            WHERE ea.event_id = ? AND e.faculty_id = ?
            ORDER BY u.firstname, u.lastname
        ''', (event_id, faculty['faculty_id'])).fetchall()
    
    conn.close()
    return jsonify([dict(student) for student in students])

@app.route('/faculty/students/edit', methods=['POST'])
def faculty_edit_student():
    if 'user_id' not in session or session['role'] != 'faculty':
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        firstname = data.get('firstname')
        lastname = data.get('lastname')
        year_level = data.get('year_level')
        is_active = data.get('is_active')
        
        if not all([user_id, firstname, lastname]):
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        conn = get_db_connection()
        
        # Verify faculty has access to this student (through classes/events)
        faculty = conn.execute('''
            SELECT f.faculty_id FROM faculty f 
            JOIN user u ON f.user_id = u.user_id 
            WHERE u.user_id = ?
        ''', (session['user_id'],)).fetchone()
        
        if not faculty:
            conn.close()
            return jsonify({'success': False, 'message': 'Faculty record not found'}), 404
        
        # Check if student is in faculty's classes or events
        student_access = conn.execute('''
            SELECT 1 FROM student_class sc
            JOIN class c ON sc.class_id = c.class_id
            JOIN student s ON sc.student_id = s.student_id
            WHERE s.user_id = ? AND c.faculty_id = ?
            UNION
            SELECT 1 FROM event_attendance ea
            JOIN event e ON ea.event_id = e.event_id
            WHERE ea.user_id = ? AND e.faculty_id = ?
        ''', (user_id, faculty['faculty_id'], user_id, faculty['faculty_id'])).fetchone()
        
        if not student_access:
            conn.close()
            return jsonify({'success': False, 'message': 'You do not have permission to edit this student'}), 403
        
        # Update user information
        conn.execute('''
            UPDATE user SET firstname = ?, lastname = ?, is_active = ?
            WHERE user_id = ?
        ''', (firstname, lastname, is_active, user_id))
        
        # Update student information if year_level is provided
        if year_level:
            conn.execute('''
                UPDATE student SET year_level = ?
                WHERE user_id = ?
            ''', (year_level, user_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Student updated successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error updating student: {str(e)}'}), 500

@app.route('/faculty/students/reset-password', methods=['POST'])
def faculty_reset_student_password():
    if 'user_id' not in session or session['role'] != 'faculty':
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        new_password = data.get('new_password')
        
        if not all([user_id, new_password]):
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        conn = get_db_connection()
        
        # Verify faculty has access to this student (through classes/events)
        faculty = conn.execute('''
            SELECT f.faculty_id FROM faculty f 
            JOIN user u ON f.user_id = u.user_id 
            WHERE u.user_id = ?
        ''', (session['user_id'],)).fetchone()
        
        if not faculty:
            conn.close()
            return jsonify({'success': False, 'message': 'Faculty record not found'}), 404
        
        # Check if student is in faculty's classes or events
        student_access = conn.execute('''
            SELECT 1 FROM student_class sc
            JOIN class c ON sc.class_id = c.class_id
            JOIN student s ON sc.student_id = s.student_id
            WHERE s.user_id = ? AND c.faculty_id = ?
            UNION
            SELECT 1 FROM event_attendance ea
            JOIN event e ON ea.event_id = e.event_id
            WHERE ea.user_id = ? AND e.faculty_id = ?
        ''', (user_id, faculty['faculty_id'], user_id, faculty['faculty_id'])).fetchone()
        
        if not student_access:
            conn.close()
            return jsonify({'success': False, 'message': 'You do not have permission to reset this student\'s password'}), 403
        
        # Update password
        conn.execute('UPDATE user SET password = ? WHERE user_id = ?', (new_password, user_id))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Password reset successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error resetting password: {str(e)}'}), 500

# Faculty My Classes
@app.route('/my_classes')
def faculty_my_classes():
    if 'user_id' not in session or session['role'] != 'faculty':
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    
    # Get faculty info
    faculty = conn.execute('''
        SELECT f.faculty_id FROM faculty f 
        JOIN user u ON f.user_id = u.user_id 
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    if not faculty:
        conn.close()
        flash('Faculty record not found', 'error')
        return redirect(url_for('faculty_dashboard'))
    
    # Get classes assigned to this faculty with student counts
    classes = conn.execute('''
        SELECT c.*, 
               COUNT(DISTINCT sc.student_id) as student_count,
               GROUP_CONCAT(DISTINCT d.day_name) as days
        FROM class c
        LEFT JOIN student_class sc ON c.class_id = sc.class_id
        LEFT JOIN class_days cd ON c.class_id = cd.class_id
        LEFT JOIN days d ON cd.day_id = d.day_id
        WHERE c.faculty_id = ?
        GROUP BY c.class_id
        ORDER BY c.class_name
    ''', (faculty['faculty_id'],)).fetchall()
    
    # Get events assigned to this faculty with attendee counts
    events = conn.execute('''
        SELECT e.*, 
               COUNT(DISTINCT ea.user_id) as attendee_count
        FROM event e
        LEFT JOIN event_attendance ea ON e.event_id = ea.event_id
        WHERE e.faculty_id = ?
        GROUP BY e.event_id
        ORDER BY e.event_date DESC
    ''', (faculty['faculty_id'],)).fetchall()
    
    # Calculate total students across all classes
    total_students = conn.execute('''
        SELECT COUNT(DISTINCT sc.student_id) as total
        FROM class c
        JOIN student_class sc ON c.class_id = sc.class_id
        WHERE c.faculty_id = ?
    ''', (faculty['faculty_id'],)).fetchone()
    
    conn.close()
    return render_template('faculty/faculty_my_classes.html', 
                         classes=classes, 
                         events=events, 
                         total_students=total_students['total'] if total_students else 0)

@app.route('/faculty/class-details/<type>/<int:id>')
def faculty_class_details(type, id):
    if 'user_id' not in session or session['role'] != 'faculty':
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = get_db_connection()
    
    # Get faculty info
    faculty = conn.execute('''
        SELECT f.faculty_id FROM faculty f 
        JOIN user u ON f.user_id = u.user_id 
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    if not faculty:
        conn.close()
        return jsonify({'error': 'Faculty record not found'}), 404
    
    if type == 'class':
        # Get class details
        class_info = conn.execute('''
            SELECT c.*, 
                   GROUP_CONCAT(DISTINCT d.day_name) as days
            FROM class c
            LEFT JOIN class_days cd ON c.class_id = cd.class_id
            LEFT JOIN days d ON cd.day_id = d.day_id
            WHERE c.class_id = ? AND c.faculty_id = ?
            GROUP BY c.class_id
        ''', (id, faculty['faculty_id'])).fetchone()
        
        if not class_info:
            conn.close()
            return jsonify({'error': 'Class not found or access denied'}), 404
        
        # Get enrolled students
        students = conn.execute('''
            SELECT u.idno, u.firstname, u.lastname, s.year_level, c.course_name
            FROM student_class sc
            JOIN student s ON sc.student_id = s.student_id
            JOIN user u ON s.user_id = u.user_id
            LEFT JOIN course c ON s.course_id = c.course_id
            WHERE sc.class_id = ?
            ORDER BY u.firstname, u.lastname
        ''', (id,)).fetchall()
        
        conn.close()
        return jsonify({
            'class_name': class_info['class_name'],
            'edpcode': class_info['edpcode'],
            'start_time': class_info['start_time'],
            'end_time': class_info['end_time'],
            'room': class_info['room'],
            'days': class_info['days'],
            'students': [dict(student) for student in students]
        })
        
    elif type == 'event':
        # Get event details
        event_info = conn.execute('''
            SELECT * FROM event 
            WHERE event_id = ? AND faculty_id = ?
        ''', (id, faculty['faculty_id'])).fetchone()
        
        if not event_info:
            conn.close()
            return jsonify({'error': 'Event not found or access denied'}), 404
        
        # Get event attendees
        attendees = conn.execute('''
            SELECT DISTINCT u.idno, u.firstname, u.lastname, s.year_level, c.course_name
            FROM event_attendance ea
            JOIN user u ON ea.user_id = u.user_id
            JOIN student s ON u.user_id = s.user_id
            LEFT JOIN course c ON s.course_id = c.course_id
            WHERE ea.event_id = ?
            ORDER BY u.firstname, u.lastname
        ''', (id,)).fetchall()
        
        conn.close()
        return jsonify({
            'event_name': event_info['event_name'],
            'description': event_info['description'],
            'event_date': event_info['event_date'],
            'start_time': event_info['start_time'],
            'end_time': event_info['end_time'],
            'room': event_info['room'],
            'attendees': [dict(attendee) for attendee in attendees]
        })
    
    conn.close()
    return jsonify({'error': 'Invalid type'}), 400

# Faculty Attendance
@app.route('/attendance')
def attendance():
    """Faculty attendance page - Take attendance using face recognition"""
    if 'user_id' not in session:
        flash('Please login to access attendance', 'error')
        return redirect(url_for('login'))
    
    # Allow both faculty and admin
    if session.get('role') not in ['faculty', 'admin']:
        flash('Only faculty and admin can access this page', 'error')
        return redirect(url_for('login'))
    
    # Get classes for faculty
    conn = get_db_connection()
    classes = []
    
    if session.get('role') == 'faculty':
        # Get faculty's classes
        faculty = conn.execute('''
            SELECT f.faculty_id FROM faculty f 
            JOIN user u ON f.user_id = u.user_id 
            WHERE u.user_id = ?
        ''', (session['user_id'],)).fetchone()
        
        if faculty:
            classes = conn.execute('''
                SELECT c.class_id, c.class_name, c.edpcode
                FROM class c
                WHERE c.faculty_id = ?
                ORDER BY c.class_name
            ''', (faculty['faculty_id'],)).fetchall()
    elif session.get('role') == 'admin':
        # Admin can see all classes
        classes = conn.execute('''
            SELECT c.class_id, c.class_name, c.edpcode
            FROM class c
            ORDER BY c.class_name
        ''').fetchall()
    
    conn.close()
    
    return render_template('faculty_attendance.html', classes=classes)

# Faculty Reports & Analytics
@app.route('/attendance_reports')
def faculty_reports():
    if 'user_id' not in session or session['role'] != 'faculty':
        return redirect(url_for('login'))
    return render_template('faculty/faculty_reports.html')

# Faculty export of today's attendance for a selected class
@app.route('/api/faculty/attendance/export/<fmt>')
def faculty_attendance_export(fmt):
    # Allow both faculty and admin
    if 'user_id' not in session or session.get('role') not in ['faculty', 'admin']:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        class_id = request.args.get('class_id')
        date_str = request.args.get('date')  # YYYY-MM-DD; defaults to today
        if not class_id:
            return jsonify({'error': 'Missing class_id'}), 400

        if not date_str:
            date_str = datetime.now().strftime('%Y-%m-%d')

        conn = get_db_connection()

        # Security: if faculty, ensure the class belongs to them
        if session.get('role') == 'faculty':
            owns = conn.execute('''
                SELECT 1
                FROM class c
                JOIN faculty f ON c.faculty_id = f.faculty_id
                JOIN user u ON f.user_id = u.user_id
                WHERE c.class_id = ? AND u.user_id = ?
            ''', (class_id, session['user_id'])).fetchone()
            if not owns:
                conn.close()
                return jsonify({'error': 'Forbidden'}), 403

        rows = conn.execute('''
            SELECT u.firstname, u.lastname, u.idno,
                   c.class_name, c.edpcode,
                   a.attendance_date, a.attendance_status
            FROM attendance a
            JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
            JOIN student s ON sc.student_id = s.student_id
            JOIN user u ON s.user_id = u.user_id
            JOIN class c ON sc.class_id = c.class_id
            WHERE sc.class_id = ? AND DATE(a.attendance_date) = ?
            ORDER BY a.attendance_date DESC
        ''', (class_id, date_str)).fetchall()
        conn.close()

        # Filename base
        filename_base = f"class_{class_id}_attendance_{date_str}"

        if fmt == 'csv':
            from io import StringIO
            import csv
            out = StringIO(); w = csv.writer(out)
            w.writerow(['Student Name', 'ID', 'Class', 'EDP Code', 'Date/Time', 'Status'])
            for r in rows:
                w.writerow([
                    f"{r['firstname']} {r['lastname']}", r['idno'] or '',
                    r['class_name'] or '', r['edpcode'] or '',
                    r['attendance_date'] or '', r['attendance_status'] or ''
                ])
            data = out.getvalue(); out.close()
            return app.response_class(
                data,
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename={filename_base}.csv'}
            )
        elif fmt == 'xlsx':
            from io import BytesIO
            try:
                from openpyxl import Workbook
            except Exception:
                return jsonify({'error': 'XLSX export not available (openpyxl missing)'}), 500
            wb = Workbook(); ws = wb.active; ws.title = 'Attendance'
            ws.append(['Student Name', 'ID', 'Class', 'EDP Code', 'Date/Time', 'Status'])
            for r in rows:
                ws.append([
                    f"{r['firstname']} {r['lastname']}", r['idno'] or '',
                    r['class_name'] or '', r['edpcode'] or '',
                    str(r['attendance_date'] or ''), r['attendance_status'] or ''
                ])
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            return app.response_class(
                bio.read(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename={filename_base}.xlsx'}
            )
        elif fmt == 'pdf':
            from io import BytesIO
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            buffer = BytesIO(); c = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter; y = height - 50
            c.setFont('Helvetica-Bold', 14)
            c.drawString(50, y, f'Attendance for Class #{class_id} on {date_str}')
            y -= 20
            c.setFont('Helvetica', 10)
            headers = ['Student', 'ID', 'Class', 'EDP', 'Date/Time', 'Status']
            c.drawString(50, y, ' | '.join(headers)); y -= 15
            for r in rows:
                line = ' | '.join([
                    f"{r['firstname']} {r['lastname']}", r['idno'] or '',
                    r['class_name'] or '', r['edpcode'] or '',
                    str(r['attendance_date'] or ''), r['attendance_status'] or ''
                ])
                # Wrap long lines
                for segment in [line[i:i+95] for i in range(0, len(line), 95)]:
                    c.drawString(50, y, segment); y -= 12
                    if y < 50:
                        c.showPage(); y = height - 50; c.setFont('Helvetica', 10)
            c.showPage(); c.save(); pdf = buffer.getvalue(); buffer.close()
            return app.response_class(
                pdf,
                mimetype='application/pdf',
                headers={'Content-Disposition': f'attachment; filename={filename_base}.pdf'}
            )
        else:
            return jsonify({'error': 'Unsupported format'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
@app.route('/api/faculty/reports/summary')
def api_faculty_reports_summary():
    if 'user_id' not in session or session['role'] != 'faculty':
        return jsonify([]), 401
    start = request.args.get('start')
    end = request.args.get('end')
    if not start or not end:
        today = datetime.now().strftime('%Y-%m-%d')
        start = today
        end = today
    conn = get_db_connection()
    faculty = conn.execute('''
        SELECT f.faculty_id FROM faculty f JOIN user u ON f.user_id = u.user_id
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    if not faculty:
        conn.close()
        return jsonify([])
    rows = conn.execute('''
        SELECT c.class_name, c.edpcode,
               COUNT(a.attendance_id) AS present_count,
               COUNT(DISTINCT sc.student_id) AS unique_students
        FROM class c
        JOIN student_class sc ON sc.class_id = c.class_id
        LEFT JOIN attendance a ON a.studentclass_id = sc.studentclass_id
            AND DATE(a.attendance_date) BETWEEN ? AND ?
        WHERE c.faculty_id = ?
        GROUP BY c.class_id
        ORDER BY c.class_name
    ''', (start, end, faculty['faculty_id'])).fetchall()
    conn.close()
    return jsonify([{
        'class_name': r['class_name'],
        'edpcode': r['edpcode'] if 'edpcode' in r.keys() else None,
        'present_count': r['present_count'] or 0,
        'unique_students': r['unique_students'] or 0
    } for r in rows])

@app.route('/api/faculty/reports/absence-patterns')
def api_faculty_reports_absence_patterns():
    if 'user_id' not in session or session['role'] != 'faculty':
        return jsonify([]), 401
    start = request.args.get('start')
    end = request.args.get('end')
    if not start or not end:
        today = datetime.now().strftime('%Y-%m-%d')
        start = today
        end = today
    conn = get_db_connection()
    faculty = conn.execute('''
        SELECT f.faculty_id FROM faculty f JOIN user u ON f.user_id = u.user_id
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    if not faculty:
        conn.close()
        return jsonify([])
    rows = conn.execute('''
        SELECT (u.firstname || ' ' || u.lastname) AS student_name,
               c.class_name,
               COUNT(a.attendance_id) AS present_count
        FROM class c
        JOIN student_class sc ON sc.class_id = c.class_id
        JOIN student s ON sc.student_id = s.student_id
        JOIN user u ON s.user_id = u.user_id
        LEFT JOIN attendance a ON a.studentclass_id = sc.studentclass_id
            AND DATE(a.attendance_date) BETWEEN ? AND ?
        WHERE c.faculty_id = ?
        GROUP BY sc.student_id, c.class_id
        HAVING present_count >= 0
        ORDER BY present_count ASC, student_name
        LIMIT 200
    ''', (start, end, faculty['faculty_id'])).fetchall()
    conn.close()
    return jsonify([{
        'student_name': r['student_name'],
        'class_name': r['class_name'],
        'present_count': r['present_count'] or 0
    } for r in rows])

@app.route('/api/faculty/reports/monthly')
def api_faculty_reports_monthly():
    if 'user_id' not in session or session['role'] != 'faculty':
        return jsonify([]), 401
    year = request.args.get('year', datetime.now().strftime('%Y'))
    conn = get_db_connection()
    faculty = conn.execute('''
        SELECT f.faculty_id FROM faculty f JOIN user u ON f.user_id = u.user_id
        WHERE u.user_id = ?
    ''', (session['user_id'],)).fetchone()
    if not faculty:
        conn.close()
        return jsonify([])
    # Initialize months 1..12 to 0
    month_counts = {str(m).zfill(2): 0 for m in range(1, 13)}
    rows = conn.execute('''
        SELECT strftime('%m', a.attendance_date) AS month,
               COUNT(a.attendance_id) AS present_count
        FROM attendance a
        JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
        JOIN class c ON sc.class_id = c.class_id
        WHERE c.faculty_id = ? AND strftime('%Y', a.attendance_date) = ?
        GROUP BY strftime('%m', a.attendance_date)
    ''', (faculty['faculty_id'], str(year))).fetchall()
    conn.close()
    for r in rows:
        month_counts[r['month']] = r['present_count'] or 0
    # Map months to labels
    month_names = {
        '01': 'Jan','02': 'Feb','03': 'Mar','04': 'Apr','05': 'May','06': 'Jun',
        '07': 'Jul','08': 'Aug','09': 'Sep','10': 'Oct','11': 'Nov','12': 'Dec'
    }
    return jsonify([{ 'month': month_names[m], 'present_count': month_counts[m] } for m in sorted(month_counts.keys())])

@app.route('/attendance_reports/export/<fmt>')
def faculty_reports_export(fmt):
    if 'user_id' not in session or session['role'] != 'faculty':
        return redirect(url_for('login'))
    
    try:
        start = request.args.get('start')
        end = request.args.get('end')
        if not start or not end:
            today = datetime.now().strftime('%Y-%m-%d')
            start = today
            end = today
        
        # Fetch datasets using the same queries
        conn = get_db_connection()
        faculty = conn.execute('''
            SELECT f.faculty_id FROM faculty f JOIN user u ON f.user_id = u.user_id
            WHERE u.user_id = ?
        ''', (session['user_id'],)).fetchone()
        
        if not faculty:
            conn.close()
            return jsonify({'error': 'No faculty record found'}), 404
        
        # Get class attendance summaries
        summary = conn.execute('''
            SELECT c.class_name, c.edpcode,
                   COUNT(a.attendance_id) AS present_count,
                   COUNT(DISTINCT sc.student_id) AS unique_students
            FROM class c
            JOIN student_class sc ON sc.class_id = c.class_id
            LEFT JOIN attendance a ON a.studentclass_id = sc.studentclass_id
                AND DATE(a.attendance_date) BETWEEN ? AND ?
            WHERE c.faculty_id = ?
            GROUP BY c.class_id
            ORDER BY c.class_name
        ''', (start, end, faculty['faculty_id'])).fetchall()
        
        # Get absence patterns
        absence = conn.execute('''
            SELECT (u.firstname || ' ' || u.lastname) AS student_name,
                   c.class_name,
                   COUNT(a.attendance_id) AS present_count
            FROM class c
            JOIN student_class sc ON sc.class_id = c.class_id
            JOIN student s ON sc.student_id = s.student_id
            JOIN user u ON s.user_id = u.user_id
            LEFT JOIN attendance a ON a.studentclass_id = sc.studentclass_id
                AND DATE(a.attendance_date) BETWEEN ? AND ?
            WHERE c.faculty_id = ?
            GROUP BY sc.student_id, c.class_id
            ORDER BY present_count ASC, student_name
        ''', (start, end, faculty['faculty_id'])).fetchall()
        
        # Get monthly attendance data
        monthly = conn.execute('''
            SELECT strftime('%Y', a.attendance_date) AS year,
                   strftime('%m', a.attendance_date) AS month,
                   COUNT(a.attendance_id) AS present_count
            FROM attendance a
            JOIN student_class sc ON a.studentclass_id = sc.studentclass_id
            JOIN class c ON sc.class_id = c.class_id
            WHERE c.faculty_id = ?
            GROUP BY strftime('%Y', a.attendance_date), strftime('%m', a.attendance_date)
            ORDER BY year, month
        ''', (faculty['faculty_id'],)).fetchall()
        
        conn.close()

        if fmt == 'csv':
            from io import StringIO
            import csv
            output = StringIO()
            writer = csv.writer(output)
            
            # Header
            writer.writerow([f"Faculty Reports & Analytics ({start} to {end})"])
            writer.writerow([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])
            
            # Class Attendance Summaries
            writer.writerow(['Class Attendance Summaries'])
            writer.writerow(['Class Name', 'EDP Code', 'Present Count', 'Unique Students'])
            for r in summary:
                writer.writerow([
                    r['class_name'] or '', 
                    r['edpcode'] or '', 
                    r['present_count'] or 0, 
                    r['unique_students'] or 0
                ])
            
            writer.writerow([])
            
            # Absence Patterns
            writer.writerow(['Absence Patterns (by low presence)'])
            writer.writerow(['Student Name', 'Class Name', 'Present Count'])
            for r in absence:
                writer.writerow([
                    r['student_name'] or '', 
                    r['class_name'] or '', 
                    r['present_count'] or 0
                ])
            
            writer.writerow([])
            
            # Monthly Attendance
            writer.writerow(['Monthly Attendance'])
            writer.writerow(['Year', 'Month', 'Present Count'])
            for r in monthly:
                writer.writerow([
                    r['year'] or '', 
                    r['month'] or '', 
                    r['present_count'] or 0
                ])
            
            csv_data = output.getvalue()
            output.close()
            
            return app.response_class(
                csv_data,
                mimetype='text/csv',
                headers={
                    'Content-Disposition': f'attachment; filename=faculty_reports_{start}_to_{end}.csv',
                    'Content-Type': 'text/csv; charset=utf-8'
                }
            )
            
        elif fmt == 'xlsx':
            try:
                from io import BytesIO
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                
                # Remove default sheet and create new ones
                wb.remove(wb.active)
                
                # Summary sheet
                ws1 = wb.create_sheet('Class Summaries')
                ws1.append([f'Faculty Reports & Analytics ({start} to {end})'])
                ws1.append([f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
                ws1.append([])
                ws1.append(['Class Name', 'EDP Code', 'Present Count', 'Unique Students'])
                
                # Style header row
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                for cell in ws1[4]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for r in summary:
                    ws1.append([
                        r['class_name'] or '', 
                        r['edpcode'] or '', 
                        r['present_count'] or 0, 
                        r['unique_students'] or 0
                    ])
                
                # Absence patterns sheet
                ws2 = wb.create_sheet('Absence Patterns')
                ws2.append([f'Faculty Reports & Analytics ({start} to {end})'])
                ws2.append([f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
                ws2.append([])
                ws2.append(['Student Name', 'Class Name', 'Present Count'])
                
                # Style header row
                for cell in ws2[4]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for r in absence:
                    ws2.append([
                        r['student_name'] or '', 
                        r['class_name'] or '', 
                        r['present_count'] or 0
                    ])
                
                # Monthly attendance sheet
                ws3 = wb.create_sheet('Monthly Attendance')
                ws3.append([f'Faculty Reports & Analytics ({start} to {end})'])
                ws3.append([f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
                ws3.append([])
                ws3.append(['Year', 'Month', 'Present Count'])
                
                # Style header row
                for cell in ws3[4]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                for r in monthly:
                    ws3.append([
                        r['year'] or '', 
                        r['month'] or '', 
                        r['present_count'] or 0
                    ])
                
                # Auto-adjust column widths
                for ws in [ws1, ws2, ws3]:
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                stream = BytesIO()
                wb.save(stream)
                stream.seek(0)
                
                return app.response_class(
                    stream.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=faculty_reports_{start}_to_{end}.xlsx'}
                )
                
            except ImportError as e:
                return jsonify({'error': f'Excel export requires openpyxl. Error: {str(e)}'}), 500
            except Exception as e:
                return jsonify({'error': f'Excel export failed: {str(e)}'}), 500
                
        elif fmt == 'pdf':
            try:
                from io import BytesIO
                from reportlab.lib.pagesizes import letter
                from reportlab.pdfgen import canvas
                from reportlab.lib.units import inch
                from reportlab.lib import colors
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_LEFT
                
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter)
                styles = getSampleStyleSheet()
                story = []
                
                # Title
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    spaceAfter=30,
                    alignment=TA_CENTER
                )
                story.append(Paragraph(f"Faculty Reports & Analytics ({start} to {end})", title_style))
                story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
                story.append(Spacer(1, 20))
                
                # Class Attendance Summaries
                story.append(Paragraph("Class Attendance Summaries", styles['Heading2']))
                story.append(Spacer(1, 12))
                
                summary_data = [['Class Name', 'EDP Code', 'Present Count', 'Unique Students']]
                for r in summary:
                    summary_data.append([
                        r['class_name'] or '', 
                        r['edpcode'] or '', 
                        str(r['present_count'] or 0), 
                        str(r['unique_students'] or 0)
                    ])
                
                summary_table = Table(summary_data)
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(summary_table)
                story.append(Spacer(1, 20))
                
                # Absence Patterns
                story.append(Paragraph("Absence Patterns (by low presence)", styles['Heading2']))
                story.append(Spacer(1, 12))
                
                absence_data = [['Student Name', 'Class Name', 'Present Count']]
                for r in absence:
                    absence_data.append([
                        r['student_name'] or '', 
                        r['class_name'] or '', 
                        str(r['present_count'] or 0)
                    ])
                
                absence_table = Table(absence_data)
                absence_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(absence_table)
                story.append(Spacer(1, 20))
                
                # Monthly Attendance
                story.append(Paragraph("Monthly Attendance", styles['Heading2']))
                story.append(Spacer(1, 12))
                
                monthly_data = [['Year', 'Month', 'Present Count']]
                for r in monthly:
                    monthly_data.append([
                        r['year'] or '', 
                        r['month'] or '', 
                        str(r['present_count'] or 0)
                    ])
                
                monthly_table = Table(monthly_data)
                monthly_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(monthly_table)
                
                doc.build(story)
                pdf = buffer.getvalue()
                buffer.close()
                
                return app.response_class(
                    pdf,
                    mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename=faculty_reports_{start}_to_{end}.pdf'}
                )
                
            except ImportError as e:
                return jsonify({'error': f'PDF export requires reportlab. Error: {str(e)}'}), 500
            except Exception as e:
                return jsonify({'error': f'PDF export failed: {str(e)}'}), 500
        else:
            return jsonify({'error': 'Unsupported format. Supported formats: csv, xlsx, pdf'}), 400
            
    except Exception as e:
        return jsonify({'error': f'Export failed: {str(e)}'}), 500

@app.route('/api/anti-spoofing/analyze', methods=['POST'])
def api_anti_spoofing_analyze():
    """Dedicated endpoint for anti-spoofing analysis"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    filepath = None
    try:
        # Get the uploaded image
        if 'image' not in request.files:
            return jsonify({'success': False, 'message': 'No image provided'}), 400
        
        file = request.files['image']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No image selected'}), 400
        
        # Save the image temporarily
        import uuid
        safe_filename = f"antispoofing_{session['user_id']}_{uuid.uuid4().hex[:8]}.jpg"
        filepath = os.path.join('temp', safe_filename)
        os.makedirs('temp', mode=0o755, exist_ok=True)
        
        # Validate file size
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > 10 * 1024 * 1024:  # 10MB limit
            return jsonify({'success': False, 'message': 'File too large'}), 400
        
        file.save(filepath)
        
        # Check if anti-spoofing is available
        if not ANTI_SPOOFING_AVAILABLE:
            return jsonify({
                'success': False, 
                'message': 'Anti-spoofing module not available',
                'is_live': True,  # Default to allowing if module unavailable
                'confidence': 0.5
            })
        
        # Load and process the image
        image = cv2.imread(filepath)
        if image is None:
            return jsonify({'success': False, 'message': 'Could not load image'}), 400
        
        # Convert to RGB
        rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        
        # Detect faces
        if FACE_RECOGNITION_AVAILABLE:
            face_locations = face_recognition.face_locations(rgb_image, model="hog")
            face_landmarks = face_recognition.face_landmarks(rgb_image, face_locations)
        else:
            return jsonify({'success': False, 'message': 'Face detection not available'}), 400
        
        if not face_locations or not face_landmarks:
            return jsonify({'success': False, 'message': 'No face detected'}), 400
        
        # Perform anti-spoofing analysis
        anti_spoofing_result = anti_spoofing_detector.comprehensive_anti_spoofing_check(
            rgb_image, face_landmarks[0], face_locations[0]
        )
        
        return jsonify({
            'success': True,
            'is_live': anti_spoofing_result['is_live'],
            'confidence': anti_spoofing_result['confidence'],
            'details': anti_spoofing_result['details'],
            'checks': anti_spoofing_result['checks']
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Analysis error: {str(e)}'}), 500
    finally:
        # Clean up temp file
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except Exception as e:
                print(f"Warning: Failed to clean up temp file {filepath}: {e}")

@app.route('/api/anti-spoofing/reset', methods=['POST'])
def api_anti_spoofing_reset():
    """Reset anti-spoofing detector state"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        if ANTI_SPOOFING_AVAILABLE:
            anti_spoofing_detector.reset_state()
            return jsonify({'success': True, 'message': 'Anti-spoofing state reset'})
        else:
            return jsonify({'success': False, 'message': 'Anti-spoofing not available'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Reset error: {str(e)}'}), 500

# ==================== SYSTEM CONFIGURATION MODULE ====================

@app.route('/settings')
def admin_settings():
    """Render admin settings page"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    return render_template('admin_settings.html')

@app.route('/api/settings/all', methods=['GET'])
def api_get_all_settings():
    """Get all system settings"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        conn = get_db_connection()
        settings = conn.execute('SELECT setting_key, setting_value FROM system_settings').fetchall()
        conn.close()
        
        settings_dict = {row['setting_key']: row['setting_value'] for row in settings}
        
        return jsonify({'success': True, 'settings': settings_dict})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/settings/save', methods=['POST'])
def api_save_settings():
    """Save system settings"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        settings = request.get_json()
        conn = get_db_connection()
        
        for key, value in settings.items():
            conn.execute('''
                UPDATE system_settings 
                SET setting_value = ?, updated_at = CURRENT_TIMESTAMP, updated_by = ?
                WHERE setting_key = ?
            ''', (str(value), session['user_id'], key))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Settings saved successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/settings/test-email', methods=['POST'])
def api_test_email():
    """Test email configuration"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        # Get email settings from database
        conn = get_db_connection()
        settings = conn.execute('''
            SELECT setting_key, setting_value 
            FROM system_settings 
            WHERE setting_type = 'email'
        ''').fetchall()
        conn.close()
        
        email_settings = {row['setting_key']: row['setting_value'] for row in settings}
        
        # Check if email is enabled
        if email_settings.get('email_enabled') != 'true':
            return jsonify({'success': False, 'error': 'Email notifications are disabled'})
        
        # Validate required fields
        required_fields = ['smtp_server', 'smtp_port', 'smtp_username', 'smtp_password', 'email_from']
        missing_fields = [f for f in required_fields if not email_settings.get(f)]
        
        if missing_fields:
            return jsonify({'success': False, 'error': f'Missing required fields: {", ".join(missing_fields)}'})
        
        # Import email libraries
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        # Create test email
        msg = MIMEMultipart()
        msg['From'] = email_settings['email_from']
        msg['To'] = email_settings['email_from']  # Send to self for testing
        msg['Subject'] = 'FaceCheck - Test Email'
        
        body = '''
        <html>
        <body>
            <h2>Test Email Successful!</h2>
            <p>This is a test email from your FaceCheck system.</p>
            <p>Your email configuration is working correctly.</p>
            <p><strong>System:</strong> FaceCheck Attendance System</p>
            <p><strong>Time:</strong> {}</p>
        </body>
        </html>
        '''.format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        msg.attach(MIMEText(body, 'html'))
        
        # Send email
        server = smtplib.SMTP(email_settings['smtp_server'], int(email_settings['smtp_port']))
        server.starttls()
        server.login(email_settings['smtp_username'], email_settings['smtp_password'])
        server.send_message(msg)
        server.quit()
        
        return jsonify({'success': True, 'message': 'Test email sent successfully'})
        
    except smtplib.SMTPAuthenticationError:
        return jsonify({'success': False, 'error': 'SMTP authentication failed. Check username and password.'})
    except smtplib.SMTPException as e:
        return jsonify({'success': False, 'error': f'SMTP error: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/settings/backup-now', methods=['POST'])
def api_backup_now():
    """Create database backup"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        import shutil
        from datetime import datetime
        
        # Create backups directory if not exists
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # Generate backup filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'facecheck_backup_{timestamp}.db'
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # Copy database file
        shutil.copy2('facecheck.db', backup_path)
        
        # Log the backup
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO system_settings (setting_key, setting_value, setting_type, description)
            VALUES (?, ?, 'backup', ?)
        ''', (f'backup_{timestamp}', backup_filename, f'Database backup created at {datetime.now()}'))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': 'Backup created successfully',
            'filename': backup_filename
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/settings/optimize-db', methods=['POST'])
def api_optimize_db():
    """Optimize database"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        conn = get_db_connection()
        
        # Run VACUUM to optimize database
        conn.execute('VACUUM')
        
        # Analyze tables for query optimization
        conn.execute('ANALYZE')
        
        conn.close()
        
        return jsonify({'success': True, 'message': 'Database optimized successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/logs')
def admin_logs():
    """View activity logs (placeholder for future implementation)"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    # This is a placeholder - you can implement full logging later
    flash('Activity logs feature coming soon!', 'info')
    return redirect(url_for('admin_settings'))

# Session check endpoint
@app.route('/api/check-session')
def check_session():
    """Check if user session is valid"""
    if 'user_id' in session:
        return jsonify({
            'valid': True,
            'user_id': session['user_id'],
            'role': session.get('role'),
            'name': f"{session.get('firstname', '')} {session.get('lastname', '')}".strip()
        })
    else:
        return jsonify({
            'valid': False,
            'message': 'No active session'
        }), 401

# Error handlers
@app.errorhandler(404)
def not_found(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

if __name__ == '__main__':
    # Create database if it doesn't exist
    if not os.path.exists('facecheck.db'):
        from db import create_database
        create_database()
    
    app.run(debug=True)